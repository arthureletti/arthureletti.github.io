"""
╔══════════════════════════════════════════════════════════════════╗
║  MYCOCAPRIN : Suivi des analyses PCR mycoplasmes caprins          ║
║  Outil réalisé par les étudiants de BUT Science des Données        ║
║  Client : Laboratoire Qualyse · Départements 79 & 86              ║
║  Équipe : Manon · Louis · Mathéo · Arthur                         ║
╚══════════════════════════════════════════════════════════════════╝

LANCEMENT :  streamlit run app.py
PRÉREQUIS :  pip install streamlit pandas plotly openpyxl xlrd
"""

import sys
import os
import html
import tempfile
from pathlib import Path

# ── Accès aux modules backend (script/) ──────────────────────────────────────
_SCRIPT = Path(__file__).parent.parent / "script"
if str(_SCRIPT) not in sys.path:
    sys.path.insert(0, str(_SCRIPT))

from code_import import ImportDonnees
from code_BD import DataBase_gestion
from code_backup import (marquer_changement, creer_backup,
                         lister_backups, restaurer_backup)
from cartographie import afficher_cartographie, classes_par_session, afficher_changement_classe


import streamlit as st
import pandas as pd
import sqlite3
import plotly.express as px
from datetime import datetime
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Libellés et couleurs des 9 classes de cheptels (règles métier Qualyse)
LIBELLES_CLASSES = {
    "A": "Toujours négatif", "B": "Toujours positif", "C": "Nouveau positif",
    "D": "Fort positif récurrent", "E": "Autre positif récurrent",
    "F": "Fort positif ponctuel", "G": "Autre positif ponctuel",
    "H": "Ancien positif", "I": "Statut non défini",
}
COULEURS_CLASSES = {
    # Palette sûre pour les troubles de la vision des couleurs (axe bleu -> orange,
    # pas de rouge/vert). Du bleu « sain » (A) au brun-orangé « préoccupant » (B).
    "A": "#1A6E8E", "H": "#4FA0BD", "C": "#8FC2D4", "F": "#EFD9A6",
    "G": "#F2B85C", "E": "#E8901F", "D": "#CC6A12", "B": "#8C3B06", "I": "#AEB6BB",
}

# ════════════════════════════════════════════════════════════════
#  CONFIGURATION GÉNÉRALE
# ════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="CapriTrack : Suivi des analyses PCR caprines",
    page_icon="assets/logo_capritrack_icone.png",
    layout="wide",
    initial_sidebar_state="expanded",
)

DB_PATH = str(Path(__file__).parent.parent / "SQL" / "data.db")

# ── CHARTE GRAPHIQUE QUALYSE ──
COULEURS = {
    "bleu_marine": "#0D4F66",
    "turquoise":   "#2BA8B8",
    "vert_sauge":  "#9DBF8E",
    "orange":      "#F39200",
    "rouge":       "#C0504D",
    "gris_clair":  "#F0F4F5",
}

# ════════════════════════════════════════════════════════════════
#  STYLE CSS (charte Qualyse)
# ════════════════════════════════════════════════════════════════
st.markdown(f"""
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
<style>
    /* Forcer le mode clair quel que soit le thème du navigateur */
    .stApp {{ background-color: #FFFFFF !important; }}
    .stApp, .stApp p, .stApp span, .stApp label, .stApp div {{ color: #262730; }}
    [data-testid="stSidebar"] {{ background-color: #F0F4F5 !important; }}

    /* Selectbox, multiselect, text input : fond blanc */
    [data-baseweb="select"] > div,
    [data-baseweb="input"] > div,
    .stSelectbox div[data-baseweb="select"] > div,
    .stMultiSelect div[data-baseweb="select"] > div,
    .stTextInput div[data-baseweb="input"] {{
        background-color: #FFFFFF !important;
        color: #262730 !important;
        border-color: #CCCCCC !important;
    }}
    [data-baseweb="select"] span, [data-baseweb="input"] input {{ color: #262730 !important; }}
    /* Menu déroulant ouvert */
    [data-baseweb="popover"], [data-baseweb="menu"], [role="listbox"] {{
        background-color: #FFFFFF !important;
    }}
    [role="option"] {{ background-color: #FFFFFF !important; color: #262730 !important; }}
    [role="option"]:hover {{ background-color: #F0F4F5 !important; }}

    /* Tags sélectionnés dans les multiselect : fond turquoise, texte blanc lisible */
    .stMultiSelect [data-baseweb="tag"] {{
        background-color: #2BA8B8 !important;
        border-radius: 6px !important;
    }}
    .stMultiSelect [data-baseweb="tag"] span {{ color: #FFFFFF !important; }}
    .stMultiSelect [data-baseweb="tag"] svg {{ fill: #FFFFFF !important; }}

    /* Tableau de données (st.dataframe / data_editor) : fond clair */
    [data-testid="stDataFrame"], [data-testid="stDataFrameResizable"],
    [data-testid="stDataEditor"], [data-testid="stDataEditorResizable"],
    [data-testid="stTable"] {{ background-color: #FFFFFF !important; }}
    [data-testid="stDataFrame"] [role="columnheader"],
    [data-testid="stDataEditor"] [role="columnheader"] {{
        background-color: #0D4F66 !important; color: #FFFFFF !important;
    }}
    [data-testid="stDataFrame"] [role="gridcell"],
    [data-testid="stDataEditor"] [role="gridcell"] {{
        background-color: #FFFFFF !important; color: #262730 !important;
    }}
    /* Variables internes de glide-data-grid (le canvas du data_editor) */
    [data-testid="stDataEditor"] > div, [data-testid="stDataFrame"] > div {{
        --gdg-bg-cell: #FFFFFF;
        --gdg-bg-cell-medium: #F7F9FA;
        --gdg-bg-header: #0D4F66;
        --gdg-bg-header-has-focus: #0a3e50;
        --gdg-bg-header-hovered: #0a3e50;
        --gdg-text-dark: #262730;
        --gdg-text-medium: #404550;
        --gdg-text-header: #FFFFFF;
        --gdg-text-header-selected: #FFFFFF;
        --gdg-border-color: #E0E0E0;
        --gdg-horizontal-border-color: #E0E0E0;
        --gdg-bg-bubble: #FFFFFF;
        --gdg-cell-horizontal-padding: 8px;
    }}
    /* Vignette du fichier chargé (file_uploader) */
    [data-testid="stFileUploaderFile"],
    [data-testid="stFileUploaderFileName"] {{
        background-color: #FFFFFF !important; color: #262730 !important;
    }}
    [data-testid="stFileUploaderFile"] * {{ color: #262730 !important; }}
    /* Champs texte : fond blanc + curseur visible */
    .stTextInput input, [data-baseweb="input"] input {{
        background-color: #FFFFFF !important;
        color: #262730 !important;
        caret-color: #0D4F66 !important;
    }}
    .stTextInput input::placeholder {{ color: #8a9aa5 !important; }}
    /* Zone de téléchargement de fichier (file_uploader) */
    [data-testid="stFileUploader"],
    [data-testid="stFileUploader"] section,
    [data-testid="stFileUploaderDropzone"],
    [data-testid="stFileUploaderDropzoneInstructions"] {{
        background-color: #F0F4F5 !important; color: #262730 !important;
    }}
    /* Rendre la vraie dropzone grande et attrayante (bordure pointillée turquoise) */
    [data-testid="stFileUploaderDropzone"] {{
        border: 2px dashed #2BA8B8 !important;
        border-radius: 12px !important;
        padding: 2rem 1.5rem !important;
        min-height: 150px !important;
        flex-direction: column !important;
        align-items: center !important;
        justify-content: center !important;
        text-align: center !important;
    }}
    /* Icône nuage au-dessus de la zone */
    [data-testid="stFileUploaderDropzone"]::before {{
        content: "\\F296";
        font-family: "bootstrap-icons" !important;
        font-size: 2.6rem;
        color: #2BA8B8;
        display: block;
        margin-bottom: 0.5rem;
    }}
    /* Masquer le texte/icône par défaut de Streamlit (anglais) */
    [data-testid="stFileUploaderDropzoneInstructions"] > div::before {{
        content: "Glissez votre fichier Excel ici, ou cliquez pour le sélectionner";
        color: #0D4F66 !important;
        font-weight: 600;
        font-size: 1.05rem;
        display: block;
    }}
    [data-testid="stFileUploaderDropzoneInstructions"] > div span,
    [data-testid="stFileUploaderDropzoneInstructions"] > div small {{
        display: none !important;
    }}
    [data-testid="stFileUploaderDropzoneInstructions"] svg {{ display: none !important; }}
    [data-testid="stFileUploader"] section * {{ color: #262730 !important; }}
    /* Tous les boutons (upload, download, génériques) : fond blanc lisible */
    [data-testid="stFileUploader"] button,
    [data-testid="stBaseButton-secondary"],
    [data-testid="baseButton-secondary"],
    [data-testid="stDownloadButton"] button,
    [data-testid="stFileUploaderDropzone"] button,
    .stDownloadButton button {{
        background-color: #FFFFFF !important;
        color: #0D4F66 !important;
        border: 1px solid #CCCCCC !important;
    }}
    [data-testid="stFileUploader"] button *,
    [data-testid="stDownloadButton"] button * {{ color: #0D4F66 !important; }}
    [data-testid="stFileUploader"] button:hover,
    [data-testid="stDownloadButton"] button:hover {{
        background-color: #F0F4F5 !important;
        border-color: #0D4F66 !important;
    }}
    /* Metrics */
    [data-testid="stMetric"] {{ background-color: transparent !important; }}
    [data-testid="stMetricValue"], [data-testid="stMetricLabel"] {{ color: #262730 !important; }}
    /* Tabs */
    .stTabs [data-baseweb="tab-panel"] {{ background-color: #FFFFFF !important; }}

    .header-bandeau {{
        background: {COULEURS['bleu_marine']};
        padding: 0.8rem 2rem;
        border-radius: 0 0 12px 12px;
        margin: -1rem -1rem 1.5rem -1rem;
        display: flex; align-items: center; justify-content: space-between;
    }}
    .header-logo {{ color: white; font-size: 1.8rem; font-weight: 800; letter-spacing: 0.05em; }}
    .header-logo .q {{ color: {COULEURS['turquoise']}; }}
    .section-titre {{
        color: {COULEURS['bleu_marine']}; font-size: 1.5rem; font-weight: 700;
        border-bottom: 3px solid {COULEURS['orange']}; padding-bottom: 0.4rem;
        margin: 1rem 0; display: inline-block;
    }}
    .kpi-card {{
        background: {COULEURS['gris_clair']}; border-radius: 12px; padding: 1.2rem;
        text-align: center; border-top: 4px solid {COULEURS['turquoise']};
    }}
    .kpi-value {{ font-size: 2rem; font-weight: 800; color: {COULEURS['bleu_marine']}; }}
    .kpi-label {{ font-size: 0.85rem; color: #5a7a8a; font-weight: 600; }}
    .stButton > button {{
        background: {COULEURS['rouge']}; color: white; border: none;
        border-radius: 24px; padding: 0.5rem 1.5rem; font-weight: 600;
    }}
    .stButton > button:hover {{ background: #a83e3b; color: white; }}
    .etape-card {{
        background: white; border: 2px solid {COULEURS['turquoise']};
        border-radius: 12px; padding: 1.2rem; text-align: center; height: 100%;
    }}
    .etape-num {{
        background: {COULEURS['turquoise']}; color: white; width: 36px; height: 36px;
        border-radius: 50%; display: inline-flex; align-items: center;
        justify-content: center; font-weight: 800; font-size: 1.1rem; margin-bottom: 0.5rem;
    }}
    .zone-import {{
        border: 2px dashed {COULEURS['turquoise']}; border-radius: 12px;
        padding: 2rem; text-align: center; background: {COULEURS['gris_clair']};
    }}

    /* Onglets principaux plus gros et lisibles */
    .stTabs [data-baseweb="tab-list"] {{ gap: 1.5rem; }}
    .stTabs [data-baseweb="tab"] {{
        font-size: 1.15rem !important;
        font-weight: 600 !important;
        padding: 0.6rem 0.4rem !important;
        color: {COULEURS['bleu_marine']} !important;
    }}
    .stTabs [data-baseweb="tab"] p {{ font-size: 1.15rem !important; font-weight: 600 !important; }}

    /* Sous-onglets façon boutons segmentés */
    div[role="radiogroup"] {{
        gap: 0.5rem !important;
        background: {COULEURS['gris_clair']};
        padding: 0.4rem;
        border-radius: 12px;
        display: inline-flex !important;
        flex-direction: row !important;
    }}
    div[role="radiogroup"] label {{
        background: transparent;
        border-radius: 8px;
        padding: 0.5rem 1.2rem !important;
        margin: 0 !important;
        cursor: pointer;
        transition: all 0.15s;
        font-weight: 600;
        color: #5a7a8a;
        border: none;
    }}
    div[role="radiogroup"] label:hover {{
        background: rgba(43,168,184,0.12);
        color: {COULEURS['bleu_marine']};
    }}
    /* Masquer le rond du radio */
    div[role="radiogroup"] label > div:first-child {{ display: none !important; }}
    /* Onglet sélectionné */
    div[role="radiogroup"] label:has(input:checked) {{
        background: {COULEURS['bleu_marine']};
        box-shadow: 0 2px 6px rgba(13,79,102,0.25);
    }}
    div[role="radiogroup"] label:has(input:checked) p,
    div[role="radiogroup"] label:has(input:checked) div {{
        color: white !important;
        font-weight: 700 !important;
    }}

    #MainMenu, footer, header {{ visibility: hidden; }}
</style>
""", unsafe_allow_html=True)

# Traduire dynamiquement les textes anglais (Select all, Browse files) en français
import streamlit.components.v1 as components
components.html("""
<script>
function traduireTextes() {
    const doc = window.parent.document;
    // "Select all" -> "Tout sélectionner"
    doc.querySelectorAll('li, [role="option"]').forEach(function(el) {
        if (el.textContent.trim() === 'Select all') {
            el.childNodes.forEach(function(node) {
                if (node.nodeType === 3 && node.textContent.trim() === 'Select all') {
                    node.textContent = 'Tout sélectionner';
                }
            });
            el.querySelectorAll('span, div').forEach(function(s) {
                if (s.textContent.trim() === 'Select all') s.textContent = 'Tout sélectionner';
            });
        }
    });
    // Bouton "Browse files" / "Upload" -> "Parcourir"
    doc.querySelectorAll('[data-testid="stFileUploader"] button').forEach(function(btn) {
        const t = btn.textContent.trim();
        if (t === 'Browse files' || t === 'Upload') {
            btn.childNodes.forEach(function(node) {
                if (node.nodeType === 3) node.textContent = ' Parcourir';
            });
            btn.querySelectorAll('span, div, p').forEach(function(s) {
                if (s.textContent.trim() === 'Browse files' || s.textContent.trim() === 'Upload')
                    s.textContent = 'Parcourir';
            });
        }
    });
}
const observer = new MutationObserver(traduireTextes);
observer.observe(window.parent.document.body, {childList: true, subtree: true});
traduireTextes();
</script>
""", height=0)


# ════════════════════════════════════════════════════════════════
#  FONCTIONS UTILITAIRES (lecture base de données)
# ════════════════════════════════════════════════════════════════
@st.cache_data(ttl=60)
def charger_donnees():
    """Charge les résultats à plat depuis SQLite (hors lignes de mélange pur)."""
    try:
        conn = sqlite3.connect(DB_PATH)
        query = """
            SELECT e.num_ede, e.nom, e.departement, e.commune,
                   p.annee, p.numero_serie,
                   (p.annee || '_serie' || p.numero_serie) AS session_label,
                   p.date_prelevement,
                   m.libelle AS mycoplasme,
                   r.valeur_ct
            FROM resultat r
            JOIN prelevement p ON r.id_prelevement = p.id_prelevement
            JOIN exploitant e  ON p.num_ede = e.num_ede
            JOIN mycoplasme m  ON r.id_mycoplasme = m.id_mycoplasme
            ORDER BY e.num_ede, p.annee, p.numero_serie
        """
        df = pd.read_sql(query, conn)
        conn.close()
        df["statut"] = df["valeur_ct"].apply(lambda x: "Négatif" if x >= 40 else "Positif")
        # Regroupement géographique : le 16 (Charente, limitrophe) est rattaché au 79
        def _zone(dept):
            if dept in ("79", "16"):
                return "Deux-Sèvres (79)"
            elif dept == "86":
                return "Vienne (86)"
            return f"Autre ({dept})"
        df["zone"] = df["departement"].apply(_zone)
        return df
    except Exception as e:
        st.error(f"Impossible de charger la base : {e}")
        return pd.DataFrame()


@st.cache_data(ttl=60)
def charger_scores():
    """Charge la classe par cheptel (table classifier : une classe par cheptel)."""
    try:
        conn = sqlite3.connect(DB_PATH)
        query = """
            SELECT c.num_ede, e.departement, e.commune, e.nom,
                   c.id_classe AS classe, cl.libelle_classe,
                   c.nb_ct_valides, c.date_calcul
            FROM classifier c
            JOIN exploitant e ON c.num_ede = e.num_ede
            LEFT JOIN classe cl ON c.id_classe = cl.id_classe
        """
        df = pd.read_sql(query, conn)
        conn.close()
        def _zone(dept):
            if dept in ("79", "16"):
                return "Deux-Sèvres (79)"
            elif dept == "86":
                return "Vienne (86)"
            return f"Autre ({dept})"
        df["zone"] = df["departement"].apply(_zone)
        return df
    except Exception:
        return pd.DataFrame()


def couleur_ct(val):
    """Mise en forme conditionnelle d'une cellule Ct."""
    if pd.isna(val):
        return "background-color: #F0F0F0; color: #999"
    elif val >= 40:
        return "background-color: #D4EFDF; color: #0E6E3E"
    elif val >= 25:
        return "background-color: #FDEBD0; color: #9C5700"
    else:
        return "background-color: #FADBD8; color: #922B21"


def pivot_vers_html(pivot):
    """
    Construit le tableau pivot en HTML avec un style en ligne sur chaque cellule.

    On n'utilise volontairement PAS Styler.to_html() : pandas y regroupe toutes
    les cellules d'une même couleur dans une seule règle CSS dont la liste de
    sélecteurs peut dépasser 50 000 caractères. Le rendu Markdown de Streamlit
    (ou le navigateur) tronque cette liste, et les cellules listées en dernier
    (lignes du bas) perdent alors leur couleur. Un style en ligne par cellule
    supprime cette dépendance et garantit la coloration de toutes les lignes.

    La couleur de fond réutilise couleur_ct() : une seule source de vérité pour
    le code couleurs, identique à la légende et aux exports.

    Paramètres
    ----------
    pivot : pd.DataFrame
        Tableau croisé cheptels (index) x sessions (colonnes), valeurs Ct.

    Retourne
    --------
    str
        Le code HTML du tableau, prêt à être inséré dans un conteneur.
    """
    style_entete = ("background-color:#0D4F66; color:white; font-weight:600; "
                    "text-align:center; padding:6px 10px; font-size:13px; "
                    "border:1px solid #E0E0E0;")
    style_cellule = ("text-align:center; padding:6px 10px; font-size:13px; "
                     "border:1px solid #E0E0E0;")

    morceaux = ['<table style="border-collapse:collapse; width:100%;">', "<thead><tr>"]
    morceaux.append(f'<th style="{style_entete}">N° EDE</th>')
    for session in pivot.columns:
        morceaux.append(f'<th style="{style_entete}">{html.escape(str(session))}</th>')
    morceaux.append("</tr></thead><tbody>")

    for ede, ligne in pivot.iterrows():
        morceaux.append("<tr>")
        morceaux.append(f'<th style="{style_entete}">{html.escape(str(ede))}</th>')
        for valeur in ligne:
            fond = couleur_ct(valeur)
            texte = "" if pd.isna(valeur) else f"{valeur:.1f}"
            morceaux.append(f'<td style="{style_cellule} {fond}">{texte}</td>')
        morceaux.append("</tr>")
    morceaux.append("</tbody></table>")
    return "".join(morceaux)


def _styles_ct(val):
    """Retourne (couleur_fond, couleur_texte) en hex pour une valeur Ct, version Excel."""
    if pd.isna(val):
        return "F0F0F0", "999999"   # gris = NA
    elif val >= 40:
        return "D4EFDF", "0E6E3E"   # vert = négatif
    elif val >= 25:
        return "FDEBD0", "9C5700"   # orange = faible positif
    else:
        return "FADBD8", "922B21"   # rouge = fort positif


def theme_clair(fig):
    """Force un rendu clair sur un graphique Plotly (sinon il suit le dark mode du navigateur)."""
    fig.update_layout(
        template="plotly_white",
        paper_bgcolor="#FFFFFF",
        plot_bgcolor="#FFFFFF",
        font=dict(color="#262730"),
        margin=dict(t=40, b=40, l=40, r=20),
    )
    fig.update_xaxes(gridcolor="#E8E8E8", zerolinecolor="#E8E8E8",
                     color="#262730", linecolor="#CCCCCC")
    fig.update_yaxes(gridcolor="#E8E8E8", zerolinecolor="#E8E8E8",
                     color="#262730", linecolor="#CCCCCC")
    return fig


def exporter_pivots_excel(dict_pivots):
    """
    Génère un fichier Excel (en mémoire) avec une feuille par cible,
    en appliquant le code couleurs Ct sur chaque cellule de données.

    dict_pivots : dict {nom_feuille: DataFrame pivot (index = EDE, colonnes = sessions)}
    Retourne les bytes du fichier .xlsx.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for nom, pivot in dict_pivots.items():
            feuille = str(nom)[:31]
            pivot.to_excel(writer, sheet_name=feuille)
            ws = writer.sheets[feuille]

            entete_fill = PatternFill("solid", fgColor="0D4F66")
            entete_font = Font(bold=True, color="FFFFFF")
            bordure = Border(*[Side(style="thin", color="CCCCCC")] * 4)
            centre = Alignment(horizontal="center", vertical="center")

            # Ligne 1 = en-têtes (sessions), colonne A = n° EDE
            for cell in ws[1]:
                cell.fill = entete_fill
                cell.font = entete_font
                cell.alignment = centre
                cell.border = bordure

            n_lignes, n_cols = pivot.shape
            # Colonne A (n° EDE) en gras
            for r in range(2, n_lignes + 2):
                c = ws.cell(row=r, column=1)
                c.font = Font(bold=True)
                c.border = bordure

            # Cellules de données : colorées selon Ct
            for i in range(n_lignes):
                for j in range(n_cols):
                    val = pivot.iat[i, j]
                    cell = ws.cell(row=i + 2, column=j + 2)
                    fond, texte = _styles_ct(val)
                    cell.fill = PatternFill("solid", fgColor=fond)
                    cell.font = Font(color=texte, bold=(pd.notna(val) and val < 40))
                    cell.alignment = centre
                    cell.border = bordure

            # Largeur des colonnes
            ws.column_dimensions["A"].width = 14
            from openpyxl.utils import get_column_letter
            for j in range(2, n_cols + 2):
                ws.column_dimensions[get_column_letter(j)].width = 13
    return buffer.getvalue()


def exporter_pivots_pdf(dict_pivots):
    """
    Génère un PDF (en mémoire) avec une page par cible, code couleurs Ct.
    dict_pivots : dict {titre: DataFrame pivot (index = EDE, colonnes = sessions)}.
    Retourne les bytes du fichier .pdf.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=1*cm, bottomMargin=1*cm,
                            leftMargin=0.8*cm, rightMargin=0.8*cm)
    elements = []
    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle("titre", parent=styles["Title"], fontSize=16,
                                 textColor=colors.HexColor("#0D4F66"))
    sous_style = ParagraphStyle("sous", parent=styles["Normal"], fontSize=8,
                                textColor=colors.grey)

    items = list(dict_pivots.items())
    for idx, (titre, pivot) in enumerate(items):
        elements.append(Paragraph(f"Tableau de suivi : {titre}", titre_style))
        elements.append(Spacer(1, 0.3*cm))
        elements.append(Paragraph(
            f"Laboratoire Qualyse · Généré le {datetime.now().strftime('%d/%m/%Y')}",
            sous_style))
        elements.append(Spacer(1, 0.4*cm))

        entetes = ["N° EDE"] + [str(c) for c in pivot.columns]
        data = [entetes]
        for ede, row in pivot.iterrows():
            ligne = [str(ede)] + ["" if pd.isna(v) else f"{v:.1f}" for v in row]
            data.append(ligne)

        n_cols = len(entetes)
        larg_ede = 2.2*cm
        larg_session = (landscape(A4)[0] - 1.6*cm - larg_ede) / max(n_cols - 1, 1)
        col_widths = [larg_ede] + [larg_session] * (n_cols - 1)

        table = Table(data, colWidths=col_widths, repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D4F66")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for i in range(len(pivot)):
            for j in range(len(pivot.columns)):
                fond, _ = _styles_ct(pivot.iat[i, j])
                style.append(("BACKGROUND", (j + 1, i + 1), (j + 1, i + 1),
                              colors.HexColor("#" + fond)))
        table.setStyle(TableStyle(style))
        elements.append(table)

        elements.append(Spacer(1, 0.5*cm))
        legende = Table([[
            "Négatif (Ct=40)", "Faiblement positif (25<=Ct<40)",
            "Fortement positif (Ct<25)", "Pas de test (NA)"
        ]], colWidths=[5*cm]*4)
        legende.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#D4EFDF")),
            ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#FDEBD0")),
            ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#FADBD8")),
            ("BACKGROUND", (3, 0), (3, 0), colors.HexColor("#F0F0F0")),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.white),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(legende)

        if idx < len(items) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    return buffer.getvalue()


def graphiques_vers_pdf(items):
    """
    Génère un PDF (en mémoire) des graphiques, rendus avec matplotlib à partir
    des données (pas de dépendance à kaleido/Chrome). Une page par graphique.

    items = liste de (titre, DataFrame, spec) où spec contient :
        {"type": "line"|"bar"|"pie", "titre": str, "x"/"y" ou "noms"/"valeurs",
         "couleur": hex, "couleurs_cat": {categorie: hex}}
    Retourne les bytes du PDF, ou None si aucun graphique.
    """
    if not items:
        return None
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    buf = io.BytesIO()
    with PdfPages(buf) as pdf:
        for titre, df, spec in items:
            fig, ax = plt.subplots(figsize=(10, 5.6))
            typ = spec.get("type", "bar")
            coul = spec.get("couleur", "#2BA8B8")
            ccat = spec.get("couleurs_cat", {})
            if typ == "line":
                x, y = spec["x"], spec["y"]
                ax.plot(df[x].astype(str), df[y], color=coul, marker="o", linewidth=2.5)
                ax.set_xlabel(x); ax.set_ylabel(y); ax.grid(True, alpha=0.3)
            elif typ == "bar":
                x, y = spec["x"], spec["y"]
                cols = [ccat.get(str(c), coul) for c in df[x]] if ccat else coul
                ax.bar(df[x].astype(str), df[y], color=cols)
                ax.set_xlabel(x); ax.set_ylabel(y); ax.grid(True, axis="y", alpha=0.3)
            elif typ == "pie":
                noms, vals = spec["noms"], spec["valeurs"]
                cols = [ccat.get(str(nm)) for nm in df[noms]] if ccat else None
                ax.pie(df[vals], labels=df[noms].astype(str), colors=cols,
                       autopct="%1.0f%%", startangle=90)
            ax.set_title(titre, fontsize=13, color="#0D4F66", fontweight="bold")
            fig.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)
    return buf.getvalue()


def graphiques_vers_excel(items):
    """
    Génère un Excel (en mémoire) : une feuille par graphique, contenant les
    données ET un graphique Excel natif éditable, avec axes (graduations) et
    les couleurs de l'application.

    items = liste de (nom_feuille, DataFrame, spec) : même spec que pour le PDF.
    Retourne les bytes du fichier .xlsx.
    """
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.marker import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.axis import ChartLines
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for feuille, df, spec in items:
            f = str(feuille)[:31]
            df.to_excel(writer, sheet_name=f, index=False)
            ws = writer.sheets[f]
            n_lignes, n_cols = df.shape

            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="0D4F66")
                cell.font = Font(bold=True, color="FFFFFF")

            typ = spec.get("type", "bar")
            chart = {"bar": BarChart, "line": LineChart, "pie": PieChart}[typ]()
            chart.title = spec.get("titre", f)
            chart.height = 9
            chart.width = 17

            if typ in ("bar", "line"):
                xcol = list(df.columns).index(spec["x"]) + 1
                ycol = list(df.columns).index(spec["y"]) + 1
                donnees = Reference(ws, min_col=ycol, max_col=ycol,
                                    min_row=1, max_row=n_lignes + 1)
                cats = Reference(ws, min_col=xcol, min_row=2, max_row=n_lignes + 1)
                chart.add_data(donnees, titles_from_data=True)
                chart.set_categories(cats)
                # Axes visibles avec graduations + titres
                chart.x_axis.title = spec["x"]
                chart.y_axis.title = spec["y"]
                chart.x_axis.delete = False
                chart.y_axis.delete = False
                chart.y_axis.majorGridlines = ChartLines()
                chart.legend = None
                serie = chart.series[0]
                serie.graphicalProperties = GraphicalProperties(
                    solidFill=spec.get("couleur", "#2BA8B8").lstrip("#"))
                ccat = spec.get("couleurs_cat", {})
                if typ == "bar" and ccat:
                    for i, c in enumerate(df[spec["x"]]):
                        col = ccat.get(str(c))
                        if col:
                            dp = DataPoint(idx=i)
                            dp.graphicalProperties = GraphicalProperties(solidFill=col.lstrip("#"))
                            serie.data_points.append(dp)
            else:  # pie
                cv = list(df.columns).index(spec["valeurs"]) + 1
                cn = list(df.columns).index(spec["noms"]) + 1
                donnees = Reference(ws, min_col=cv, max_col=cv,
                                    min_row=1, max_row=n_lignes + 1)
                cats = Reference(ws, min_col=cn, min_row=2, max_row=n_lignes + 1)
                chart.add_data(donnees, titles_from_data=True)
                chart.set_categories(cats)
                ccat = spec.get("couleurs_cat", {})
                if ccat:
                    for i, nm in enumerate(df[spec["noms"]]):
                        col = ccat.get(str(nm))
                        if col:
                            dp = DataPoint(idx=i)
                            dp.graphicalProperties = GraphicalProperties(solidFill=col.lstrip("#"))
                            chart.series[0].data_points.append(dp)

            ws.add_chart(chart, f"{get_column_letter(n_cols + 2)}2")
            for j in range(1, n_cols + 1):
                ws.column_dimensions[get_column_letter(j)].width = 18
    return buf.getvalue()


def tableau_simple_excel(df, feuille, titre, col_couleur=None, couleurs=None):
    """
    Exporte un DataFrame simple en Excel (en-tête charte Qualyse).
    Si col_couleur est fourni, colore le fond des cellules de cette colonne
    selon le dictionnaire couleurs {valeur: hex} (ex. couleurs des classes).
    """
    from openpyxl.utils import get_column_letter
    couleurs = couleurs or {}
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        f = str(feuille)[:31]
        df.to_excel(writer, sheet_name=f, index=False)
        ws = writer.sheets[f]
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="0D4F66")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_couleur and col_couleur in df.columns:
            j = list(df.columns).index(col_couleur) + 1
            for i, val in enumerate(df[col_couleur], start=2):
                hexa = couleurs.get(str(val))
                if hexa:
                    ws.cell(row=i, column=j).fill = PatternFill(
                        "solid", fgColor=hexa.lstrip("#"))
                    ws.cell(row=i, column=j).font = Font(bold=True, color="FFFFFF")
        for j in range(1, df.shape[1] + 1):
            ws.column_dimensions[get_column_letter(j)].width = 18
    return buf.getvalue()


def tableau_simple_pdf(df, titre, col_couleur=None, couleurs=None):
    """
    Exporte un DataFrame simple en PDF (tableau, en-tête charte).
    Si col_couleur est fourni, colore le fond des cellules de cette colonne
    selon couleurs {valeur: hex}.
    """
    from reportlab.lib import colors as rlcolors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    couleurs = couleurs or {}
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.2*cm, bottomMargin=1.2*cm,
                            leftMargin=1.2*cm, rightMargin=1.2*cm)
    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle("t", parent=styles["Title"], fontSize=15,
                                 textColor=rlcolors.HexColor("#0D4F66"))
    elements = [Paragraph(titre, titre_style), Spacer(1, 0.3*cm),
                Paragraph(f"Laboratoire Qualyse · {datetime.now().strftime('%d/%m/%Y')}",
                          ParagraphStyle("s", parent=styles["Normal"], fontSize=8,
                                         textColor=rlcolors.grey)),
                Spacer(1, 0.4*cm)]
    data = [list(df.columns)] + df.astype(str).values.tolist()
    table = Table(data, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), rlcolors.HexColor("#0D4F66")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rlcolors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, rlcolors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rlcolors.white, rlcolors.HexColor("#F0F4F5")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    if col_couleur and col_couleur in df.columns:
        c = list(df.columns).index(col_couleur)
        for r, val in enumerate(df[col_couleur], start=1):
            hexa = couleurs.get(str(val))
            if hexa:
                style.append(("BACKGROUND", (c, r), (c, r), rlcolors.HexColor(hexa)))
                style.append(("TEXTCOLOR", (c, r), (c, r), rlcolors.white))
                style.append(("FONTNAME", (c, r), (c, r), "Helvetica-Bold"))
    table.setStyle(TableStyle(style))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def _exports_graphiques(prefixe, items):
    """
    Affiche les boutons d'export des graphiques d'une section.
    items = liste de (titre, DataFrame, spec). Excel = données + graphiques
    natifs ; PDF = graphiques rendus avec matplotlib (rapide, sans kaleido).
    """
    if not items:
        return
    st.markdown("#### Exporter les graphiques")
    _h = datetime.now().strftime("%Y%m%d")
    col_x, col_p = st.columns(2)
    with col_x:
        try:
            st.download_button(
                "Excel (données + graphiques)",
                data=graphiques_vers_excel(items),
                file_name=f"graphiques_{prefixe}_{_h}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key=f"xlsx_{prefixe}")
        except Exception as e:
            st.warning(f"Export Excel indisponible : {e}")
    with col_p:
        try:
            st.download_button(
                "PDF des graphiques",
                data=graphiques_vers_pdf(items),
                file_name=f"graphiques_{prefixe}_{_h}.pdf", mime="application/pdf",
                use_container_width=True, key=f"pdf_{prefixe}")
        except Exception as e:
            st.warning(f"Export PDF indisponible : {e}")


def valider_donnees_import(df, db_path):
    """
    Valide le contenu d'un DataFrame avant l'import définitif en base.

    Retourne un tuple (erreurs, avertissements) où chaque élément est une
    liste de messages clairs destinés à l'utilisateur.

    Erreurs BLOQUANTES (empêchent l'import) :
      - valeurs nulles/manquantes dans les colonnes obligatoires ;
      - n° EDE mal formé (doit faire 8 chiffres) ;
      - valeur Ct non numérique ou hors plage plausible (0 à 40).

    AVERTISSEMENTS (non bloquants) :
      - matrice absente de la base (nouvelle matrice) ;
      - n° EDE absent de la base (nouveau cheptel) ;
      - dates de prélèvement manquantes.
    """
    erreurs, avertissements = [], []

    OBLIGATOIRES = ["N EDE", "MATRICE"]

    # ── 1. Colonnes obligatoires : valeurs nulles / manquantes ───────────────
    for col in OBLIGATOIRES:
        if col not in df.columns:
            continue  # l'absence de colonne est déjà gérée par detecter_tt_col()
        nb_nuls = int(df[col].isna().sum())
        if nb_nuls > 0:
            erreurs.append(f"{nb_nuls} valeur(s) manquante(s) dans la colonne obligatoire « {col} ».")

    # ── 2. Format du n° EDE (8 chiffres) ─────────────────────────────────────
    if "N EDE" in df.columns:
        ede = df["N EDE"].dropna().astype(str)
        mal_formes = ede[~ede.str.fullmatch(r"\d{8}")]
        if len(mal_formes) > 0:
            exemples = ", ".join(mal_formes.unique()[:3])
            erreurs.append(
                f"{len(mal_formes)} n° EDE mal formé(s) (8 chiffres attendus). "
                f"Exemple(s) : {exemples}."
            )

    # ── 3. Valeurs Ct (colonnes PCR) numériques et dans la plage 0–40 ────────
    _v = ImportDonnees.__new__(ImportDonnees)
    _v.df = df
    cols_pcr = _v.detecter_PCR()["colonnes"]
    for col in cols_pcr:
        serie_num = pd.to_numeric(df[col], errors="coerce")
        nb_non_num = int(serie_num.isna().sum() - df[col].isna().sum())
        if nb_non_num > 0:
            erreurs.append(f"{nb_non_num} valeur(s) non numérique(s) dans la colonne PCR « {col} ».")
        hors_plage = serie_num.dropna()
        hors_plage = hors_plage[(hors_plage < 0) | (hors_plage > 40)]
        if len(hors_plage) > 0:
            erreurs.append(
                f"{len(hors_plage)} valeur(s) Ct hors plage (0 à 40) dans « {col} » "
                "- un Ct négatif vaut 40, un positif est < 40."
            )

    # ── 4. Avertissements : dates manquantes ─────────────────────────────────
    if "DATE PRELVMT" in df.columns:
        nb_dates_nulles = int(df["DATE PRELVMT"].isna().sum())
        if nb_dates_nulles > 0:
            avertissements.append(f"{nb_dates_nulles} date(s) de prélèvement manquante(s).")

    # Nom d'exploitant manquant : non bloquant. Le cheptel reste identifié par
    # son n° EDE ; le nom pourra être complété par la suite.
    if "NOM EXPLOITANT" in df.columns:
        nb_nuls = int(df["NOM EXPLOITANT"].isna().sum())
        nb_vides = int((df["NOM EXPLOITANT"].astype(str).str.strip() == "").sum())
        nb_sans_nom = max(nb_nuls, nb_vides)
        if nb_sans_nom > 0:
            avertissements.append(
                f"{nb_sans_nom} nom(s) d'exploitant manquant(s) (non bloquant)."
            )

    # ── 5. Avertissements : matrices / EDE inconnus en base ──────────────────
    try:
        conn = sqlite3.connect(db_path)
        matrices_base = {r[0] for r in conn.execute("SELECT libelle FROM matrice").fetchall()}
        edes_base = {r[0] for r in conn.execute("SELECT num_ede FROM exploitant").fetchall()}
        conn.close()

        if "MATRICE" in df.columns and matrices_base:
            nouvelles = set(df["MATRICE"].dropna().unique()) - matrices_base
            if nouvelles:
                avertissements.append(
                    "Nouveau(x) type(s) de prélèvement (matrice) détecté(s) : "
                    + ", ".join(sorted(str(m) for m in nouvelles))
                    + ". Ils seront ajoutés automatiquement à la liste de référence "
                    "(ex. « Lait de tank », « Buvard FTA »)."
                )
        if "N EDE" in df.columns and edes_base:
            nouveaux = set(df["N EDE"].dropna().astype(str).unique()) - edes_base
            if nouveaux:
                avertissements.append(
                    f"{len(nouveaux)} nouveau(x) cheptel(s) (EDE absent(s) de la base, seront créés)."
                )
    except Exception:
        pass  # base absente/vide : pas d'avertissement possible, non bloquant

    return erreurs, avertissements


def carte_erreurs_cellules(df, db_path):
    """
    Repère, cellule par cellule, les anomalies des données avant import.

    Retourne (carte, nb_err, nb_warn, nb_info) où `carte` est un dict
    {index_ligne: {nom_colonne: (severite, message)}} avec severite ∈
    {"err" (bloquant, rouge), "warn" (à vérifier, orange), "info" (bleu)}.

    Anomalies détectées :
      - N° EDE manquant / invalide (8 chiffres) → erreur (rouge)
      - N° EDE inconnu en base → info (bleu, nouvel exploitant créé)
      - Nom d'exploitant manquant → avertissement (orange, non bloquant)
      - Date de prélèvement manquante → non bloquant (conservée à None)
      - Valeur Ct manquante → avertissement (orange, résultat manquant)
      - Valeur Ct non numérique / hors plage 0–40 → erreur (rouge)
    """
    import re as _re
    edes_base, matrices_base = set(), set()
    try:
        _c = sqlite3.connect(db_path)
        edes_base = {str(r[0]) for r in _c.execute("SELECT num_ede FROM exploitant")}
        matrices_base = {str(r[0]) for r in _c.execute("SELECT libelle FROM matrice")}
        _c.close()
    except Exception:
        pass

    _v = ImportDonnees.__new__(ImportDonnees)
    _v.df = df
    try:
        cols_pcr = _v.detecter_PCR()["colonnes"]
    except Exception:
        cols_pcr = [c for c in df.columns if str(c).startswith("M.")]

    carte = {}

    def _add(i, col, sev, msg):
        carte.setdefault(i, {})[col] = (sev, msg)

    def _vide(v):
        return pd.isna(v) or str(v).strip() in ("", "nan", "None", "NaT", "<NA>")

    for i in df.index:
        if "N EDE" in df.columns:
            v = df.at[i, "N EDE"]
            s = "" if pd.isna(v) else str(v).strip()
            if _vide(v):
                _add(i, "N EDE", "err", "N° EDE manquant")
            elif not _re.fullmatch(r"\d{8}", s):
                _add(i, "N EDE", "err", f"N° EDE invalide : « {s} » (8 chiffres attendus)")
            elif edes_base and s not in edes_base:
                _add(i, "N EDE", "info", "Cheptel inconnu en base : un nouvel exploitant sera créé")
        # Nom d'exploitant manquant : NON bloquant. Le cheptel reste identifié
        # par son n° EDE ; on signale en orange (à vérifier), pas en rouge.
        if "NOM EXPLOITANT" in df.columns and _vide(df.at[i, "NOM EXPLOITANT"]):
            _add(i, "NOM EXPLOITANT", "warn", "Nom d'exploitant manquant (non bloquant)")
        # Date manquante : NON bloquant (conservée vide). L'année et la série
        # suffisent à situer le prélèvement dans le temps. On la signale juste
        # en bleu (information), pas en rouge.
        if "DATE PRELVMT" in df.columns and _vide(df.at[i, "DATE PRELVMT"]):
            _add(i, "DATE PRELVMT", "info",
                 "Date de prélèvement absente (non bloquant, conservée vide)")
        for col in cols_pcr:
            if col not in df.columns:
                continue
            v = df.at[i, col]
            if _vide(v):
                _add(i, col, "warn", "Résultat manquant pour cette cible")
            else:
                num = pd.to_numeric(pd.Series([v]), errors="coerce").iloc[0]
                if pd.isna(num):
                    _add(i, col, "err", f"Valeur Ct non numérique : « {v} »")
                elif num < 0 or num > 40:
                    _add(i, col, "err", f"Ct hors plage 0–40 : {num} (négatif = 40, positif < 40)")

    nb_err = sum(1 for r in carte.values() for sev, _ in r.values() if sev == "err")
    nb_warn = sum(1 for r in carte.values() for sev, _ in r.values() if sev == "warn")
    nb_info = sum(1 for r in carte.values() for sev, _ in r.values() if sev == "info")
    return carte, nb_err, nb_warn, nb_info


def html_tableau_erreurs(df, carte, max_lignes=None):
    """
    Construit un tableau HTML (lecture seule) limité aux lignes contenant une
    anomalie : les cellules fautives sont colorées (rouge = erreur bloquante,
    orange = à vérifier, bleu = nouveau cheptel) et affichent le message au
    survol. Utilisé uniquement en repli si le composant AgGrid n'est pas
    disponible. max_lignes=None affiche toutes les lignes en anomalie.
    """
    from html import escape
    cols = list(df.columns)
    lignes_pb = [i for i in df.index if i in carte]
    if not lignes_pb:
        return ""
    tronque = max_lignes is not None and len(lignes_pb) > max_lignes
    if max_lignes is not None:
        lignes_pb = lignes_pb[:max_lignes]

    th = ('<th style="position:sticky;top:0;background:#0D4F66;color:#fff;'
          'padding:6px 8px;font-size:0.8rem;white-space:nowrap;">Ligne</th>')
    th += "".join(
        f'<th style="position:sticky;top:0;background:#0D4F66;color:#fff;'
        f'padding:6px 8px;font-size:0.8rem;white-space:nowrap;">{escape(str(c))}</th>'
        for c in cols)

    corps = ""
    for i in lignes_pb:
        tds = (f'<td style="font-weight:700;color:#0D4F66;padding:5px 8px;'
               f'border-bottom:1px solid #eee;">{i + 1}</td>')
        for c in cols:
            v = df.at[i, c]
            txt = "" if pd.isna(v) else escape(str(v))
            if i in carte and c in carte[i]:
                sev, msg = carte[i][c]
                couleurs = {"err": ("#F4C6C6", "#C0392B"),
                            "warn": ("#FBE2BB", "#E08600"),
                            "info": ("#D8E6F5", "#2F6FB0")}
                bg, bord = couleurs.get(sev, ("#EEEEEE", "#999999"))
                contenu = txt if txt else "-"
                tds += (f'<td title="{escape(msg)}" style="background:{bg};'
                        f'box-shadow:inset 3px 0 0 {bord};padding:5px 8px;'
                        f'border-bottom:1px solid #eee;cursor:help;font-size:0.82rem;">'
                        f'{contenu}</td>')
            else:
                tds += (f'<td style="padding:5px 8px;border-bottom:1px solid #eee;'
                        f'font-size:0.82rem;color:#333;">{txt}</td>')
        corps += f"<tr>{tds}</tr>"

    note = ("<div style='font-size:0.78rem;color:#888;margin-top:0.4rem;'>"
            f"Affichage limité aux {max_lignes} premières lignes en anomalie.</div>"
            if tronque else "")
    return (
        "<div style='overflow:auto;max-height:430px;border:1px solid #e0e6e9;"
        "border-radius:8px;'>"
        "<table style='border-collapse:collapse;width:100%;background:#fff;'>"
        f"<thead><tr>{th}</tr></thead><tbody>{corps}</tbody></table></div>" + note
    )


def editeur_donnees_import(df_source, db_path):
    """
    Affiche UN SEUL tableau, à la fois éditable et avec coloration des
    anomalies + infobulles, grâce au composant AgGrid.

    - Cellules colorées : rouge = erreur bloquante, orange = à vérifier,
      bleu = nouveau cheptel. Le message s'affiche au survol.
    - Édition directe dans le tableau ; ajout/suppression de lignes possibles.

    Si le composant AgGrid n'est pas installé, on bascule automatiquement sur
    l'éditeur Streamlit + un tableau coloré en lecture seule (repli).

    Retourne (df_edite, nb_err, nb_warn, nb_info).
    """
    try:
        from st_aggrid import AgGrid, GridOptionsBuilder, JsCode, GridUpdateMode
    except Exception:
        return _editeur_donnees_fallback(df_source, db_path)

    import json

    # df de travail persistant : sert de base au calcul des couleurs. Réinitialisé
    # dès qu'un fichier DIFFÉRENT est chargé. La signature inclut un condensé du
    # CONTENU (et pas seulement la taille/les colonnes) pour éviter qu'un autre
    # fichier de même forme réutilise par erreur l'état précédent.
    try:
        empreinte = pd.util.hash_pandas_object(df_source, index=False).sum()
    except Exception:
        empreinte = 0
    sig = f"{len(df_source)}|{list(df_source.columns)}|{empreinte}"
    if (st.session_state.get("_edit_sig") != sig
            or "_edit_df" not in st.session_state):
        st.session_state["_edit_df"] = df_source.copy().reset_index(drop=True)
        st.session_state["_edit_sig"] = sig
        st.session_state["_edit_reload"] = True

    travail = st.session_state["_edit_df"]
    carte, nb_err, nb_warn, nb_info = carte_erreurs_cellules(travail, db_path)

    # Préparation pour la grille : TOUTES les valeurs en texte (évite l'affichage
    # « invalid number » d'AgGrid quand une colonne numérique contient des
    # cellules vides) + colonne _err cachée. Les nombres entiers sont affichés
    # sans « .0 » ; la reconversion en numérique/date se fait après l'édition.
    def _fmt_cellule(v):
        if pd.isna(v):
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)

    df_grid = travail.copy()
    for c in df_grid.columns:
        if "datetime" in str(df_grid[c].dtype):
            _d = pd.to_datetime(df_grid[c], errors="coerce")
            # Une date antérieure à 2000 est un artefact (ex. epoch 01-01-1970
            # issu d'une cellule vide) : on l'affiche comme vide.
            _d = _d.where(_d.dt.year >= 2000)
            df_grid[c] = _d.dt.strftime("%Y-%m-%d")
        df_grid[c] = df_grid[c].map(_fmt_cellule)
    # Noms de colonnes « sûrs » pour AgGrid : un point dans le nom
    # (« M. MYCOIDES CAPRI ») déclenche la notation pointée d'AgGrid, qui réécrit
    # la valeur en données imbriquées et fait disparaître la colonne au retour.
    # On retire donc le point du champ technique tout en gardant un en-tête lisible.
    reel_vers_grille = {c: str(c).replace(".", "") for c in df_grid.columns if "." in str(c)}
    grille_vers_reel = {v: k for k, v in reel_vers_grille.items()}

    df_grid["_err"] = [
        json.dumps({str(reel_vers_grille.get(col, col)): [sev, msg]
                    for col, (sev, msg) in carte.get(i, {}).items()},
                   ensure_ascii=False)
        for i in range(len(df_grid))
    ]
    df_grid["_rid"] = list(range(len(df_grid)))  # identifiant de ligne stable

    cell_style = JsCode("""
        function(p){
          try{
            var m = JSON.parse(p.data._err || '{}');
            var info = m[p.colDef.field];
            if(info){
              var s = info[0];
              if(s==='err')  return {'backgroundColor':'#F4C6C6','boxShadow':'inset 3px 0 0 #C0392B'};
              if(s==='warn') return {'backgroundColor':'#FBE2BB','boxShadow':'inset 3px 0 0 #E08600'};
              if(s==='info') return {'backgroundColor':'#D8E6F5','boxShadow':'inset 3px 0 0 #2F6FB0'};
            }
          }catch(e){}
          return null;
        }
    """)
    tooltip = JsCode("""
        function(p){
          try{
            var m = JSON.parse(p.data._err || '{}');
            var info = m[p.colDef.field];
            if(info){ return info[1]; }
          }catch(e){}
          return '';
        }
    """)

    df_grid_aff = df_grid.rename(columns=reel_vers_grille)
    gb = GridOptionsBuilder.from_dataframe(df_grid_aff)
    gb.configure_default_column(editable=True, resizable=True, sortable=True,
                                wrapHeaderText=True, autoHeaderHeight=True,
                                cellStyle=cell_style, tooltipValueGetter=tooltip)
    # En-tête lisible (nom réel avec point) malgré le champ technique sans point.
    for _safe, _reel in grille_vers_reel.items():
        gb.configure_column(_safe, headerName=_reel)
    gb.configure_column("_err", hide=True, editable=False)
    gb.configure_column("_rid", hide=True, editable=False)
    # Cases à cocher pour sélectionner des lignes à supprimer
    gb.configure_selection("multiple", use_checkbox=True, header_checkbox=True)
    grid_options = gb.build()
    grid_options["enableBrowserTooltips"] = True
    # Les noms de colonnes contiennent des points (« M. MYCOIDES CAPRI »…).
    # Sans cette option, AgGrid les interprète comme des chemins imbriqués
    # (data["M"][" MYCOIDES CAPRI"]) et la colonne apparaît vide / disparaît.
    grid_options["suppressFieldDotNotation"] = True
    # Réajuste les colonnes quand la grille change de taille (ex. passage d'une
    # fenêtre réduite à agrandie) : le tableau remplit alors toute la largeur.
    grid_options["onGridSizeChanged"] = JsCode(
        "function(p){ if(p.api){ p.api.sizeColumnsToFit(); } }")

    # Filtre : n'afficher que les lignes en anomalie (le tableau reste éditable
    # et supprimable ; les éditions sont fusionnées dans le jeu complet via _rid).
    lignes_err = set(carte.keys())
    afficher_que_err = st.checkbox(
        f"Afficher seulement les lignes en anomalie ({len(lignes_err)})",
        key="filtre_anomalies_import", value=False, disabled=not lignes_err)
    if afficher_que_err and lignes_err:
        df_display = df_grid_aff[df_grid_aff["_rid"].isin(lignes_err)].copy()
    else:
        df_display = df_grid_aff

    recharger = st.session_state.pop("_edit_reload", False)
    nonce = st.session_state.get("_grid_nonce", 0)
    # Le filtre est inclus dans le key : basculer le filtre remonte la grille.
    reponse = AgGrid(
        df_display, gridOptions=grid_options, allow_unsafe_jscode=True,
        update_mode=GridUpdateMode.VALUE_CHANGED, height=460,
        fit_columns_on_grid_load=True, reload_data=recharger,
        key=f"aggrid_import_{nonce}_{int(afficher_que_err)}",
    )

    # Fusion des éditions du sous-ensemble affiché dans le tableau COMPLET (_rid),
    # pour ne pas perdre les lignes masquées par le filtre.
    base = df_grid.drop(columns=["_err"]).copy()
    edites = pd.DataFrame(reponse["data"]) if reponse and "data" in reponse else df_display.copy()
    # Retirer les colonnes internes d'AgGrid qui peuvent fuiter (::auto_unique_id::…)
    edites = edites[[c for c in edites.columns if not str(c).startswith("::")]]
    # Revenir aux noms réels (avec points) avant toute fusion.
    if grille_vers_reel:
        edites = edites.rename(columns=grille_vers_reel)
    if "_err" in edites.columns:
        edites = edites.drop(columns=["_err"])
    if "_rid" in edites.columns and not edites.empty:
        edites["_rid"] = pd.to_numeric(edites["_rid"], errors="coerce")
        # AgGrid peut renvoyer des _rid vides (NaN), des _rid dupliqués, voire
        # une colonne en double, après sélection/ajout/suppression de lignes.
        # base.update() lève alors pandas InvalidIndexError (il exige un index
        # ET des colonnes uniques). On nettoie puis on fusionne À LA MAIN,
        # colonne par colonne sur l'intersection des _rid : aucune opération
        # ne peut plus lever InvalidIndexError.
        edites = edites.dropna(subset=["_rid"]).drop_duplicates("_rid", keep="last")
        edites = edites.loc[:, ~edites.columns.duplicated()]
        base = base.loc[:, ~base.columns.duplicated()]
        base = base.drop_duplicates("_rid", keep="last").set_index("_rid")
        e = edites.set_index("_rid")
        if not e.empty:
            communes = [c for c in e.columns if c in base.columns]
            idx = base.index.intersection(e.index)
            for c in communes:
                base.loc[idx, c] = e.loc[idx, c]
        base = base.reset_index()

    def _vers_import(df_txt):
        """Forme texte de la grille -> types corrects pour validation/import."""
        r = df_txt.drop(columns=["_rid"], errors="ignore").copy()
        r = r.replace({"": pd.NA, "nan": pd.NA, "NaT": pd.NA})
        if "DATE PRELVMT" in r.columns:
            _dt = pd.to_datetime(r["DATE PRELVMT"], errors="coerce")
            # Neutralise l'artefact epoch (01-01-1970) issu d'une cellule vide.
            _dt = _dt.where(_dt.dt.year >= 2000)
            r["DATE PRELVMT"] = _dt
        for _c in r.columns:
            if str(_c).startswith("M."):
                r[_c] = pd.to_numeric(r[_c], errors="coerce")
        return r

    out = _vers_import(base)
    # Persistance : on mémorise les éditions à chaque passage (sans rerun, donc
    # sans boucle) pour qu'elles survivent aux changements de filtre.
    st.session_state["_edit_df"] = out.reset_index(drop=True)

    # Lignes cochées (identifiées par _rid)
    sel = reponse.get("selected_rows", None) if reponse else None
    rids = []
    try:
        if sel is not None:
            if hasattr(sel, "columns"):
                if "_rid" in sel.columns:
                    rids = [int(x) for x in sel["_rid"].tolist()]
            elif isinstance(sel, list):
                rids = [int(r["_rid"]) for r in sel if r.get("_rid") is not None]
    except Exception:
        rids = []

    col_suppr, col_rev = st.columns(2)
    with col_suppr:
        if st.button(f"Supprimer la sélection ({len(rids)} ligne(s))",
                     key="suppr_lignes_import", disabled=not rids,
                     use_container_width=True):
            reste = base[~base["_rid"].isin(set(rids))]
            st.session_state["_edit_df"] = _vers_import(reste).reset_index(drop=True)
            st.session_state["_edit_reload"] = True
            st.session_state["_grid_nonce"] = nonce + 1   # vide la sélection
            st.rerun()
    with col_rev:
        # Remonte la grille pour recalculer la coloration des cellules.
        if st.button("Revérifier les anomalies", key="revanaly_import",
                     use_container_width=True):
            st.session_state["_grid_nonce"] = nonce + 1
            st.session_state["_edit_reload"] = True
            st.rerun()
    st.caption("Pour supprimer une ou plusieurs lignes : cochez-les (1re colonne), "
               "puis cliquez sur « Supprimer la sélection ».")

    # Compteurs LIVE : calculés sur les données réellement éditées (out), ils se
    # mettent donc à jour à chaque correction, même avant de rafraîchir les couleurs.
    _, nb_err, nb_warn, nb_info = carte_erreurs_cellules(out, db_path)

    return out, nb_err, nb_warn, nb_info


def _editeur_donnees_fallback(df_source, db_path):
    """
    Repli si AgGrid n'est pas disponible : éditeur Streamlit classique + tableau
    coloré en lecture seule (deux tableaux). Retourne (df_edite, e, w, info).
    """
    df_pour_edition = df_source.copy()
    for _c in df_pour_edition.columns:
        if str(df_pour_edition[_c].dtype) in ("str", "string"):
            df_pour_edition[_c] = df_pour_edition[_c].astype(object)
    df_edite = st.data_editor(
        df_pour_edition, use_container_width=True, num_rows="dynamic",
        height=420, key="editeur_import",
    )
    carte, nb_err, nb_warn, nb_info = carte_erreurs_cellules(df_edite, db_path)
    if nb_err or nb_warn or nb_info:
        st.markdown(html_tableau_erreurs(df_edite, carte), unsafe_allow_html=True)
        st.caption("Composant AgGrid non installé : pour un tableau unique éditable et "
                   "coloré, installez-le avec « pip install streamlit-aggrid ».")
    return df_edite, nb_err, nb_warn, nb_info


# ════════════════════════════════════════════════════════════════
#  HEADER / BANDEAU SUPÉRIEUR
# ════════════════════════════════════════════════════════════════
import base64 as _b64
def _logo_b64(path):
    try:
        with open(path, "rb") as f:
            return _b64.b64encode(f.read()).decode()
    except Exception:
        return ""

_icone = _logo_b64("assets/logo_capritrack_icone.png")
_icone_html = (f'<img src="data:image/png;base64,{_icone}" style="height:54px; '
               f'background:white; border-radius:8px; padding:3px;">'
               if _icone else '<i class="bi bi-droplet-fill" style="font-size:2rem; color:white;"></i>')

# Couleurs du logo CapriTrack : "Capri" bleu marine, "Track" turquoise, slogan bleu
st.markdown(f"""
<div class="header-bandeau">
    <div style="display:flex; align-items:center; gap:0.9rem;">
        {_icone_html}
        <div style="display:flex; align-items:center; gap:0.9rem;">
            <div class="header-logo" style="margin:0;">Capri<span class="q">Track</span></div>
            <div style="width:2px; height:42px; background:#2BA8B8; opacity:0.6;"></div>
            <div style="color:#9DBF8E; font-size:0.78rem; font-weight:600;
                        letter-spacing:0.08em; line-height:1.3;">
                SUIVI · ANALYSE<br>PERFORMANCE
            </div>
        </div>
    </div>
    <div style="color:#9DBF8E; font-size:0.85rem;">Laboratoire Qualyse · Départements 79 &amp; 86</div>
</div>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
#  NAVIGATION PRINCIPALE (3 onglets)
# ════════════════════════════════════════════════════════════════
onglet_accueil, onglet_import, onglet_visu, onglet_sauvegarde = st.tabs([
    ":material/home: Accueil",
    ":material/upload_file: Importer",
    ":material/bar_chart: Visualisation & Analyses",
    ":material/save: Sauvegardes",
])

# État de session partagé entre tous les onglets
if "bdd_modifiee" not in st.session_state:
    st.session_state.bdd_modifiee = False

# Popup central de confirmation d'import (s'estompe tout seul via animation CSS)
if st.session_state.get("import_succes_msg"):
    _msg_popup = st.session_state.import_succes_msg
    del st.session_state.import_succes_msg
    st.markdown(f"""
    <style>
    @keyframes capri_fade {{
        0%   {{ opacity: 0; transform: translate(-50%, -45%) scale(0.96); }}
        12%  {{ opacity: 1; transform: translate(-50%, -50%) scale(1); }}
        70%  {{ opacity: 1; transform: translate(-50%, -50%) scale(1); }}
        100% {{ opacity: 0; transform: translate(-50%, -50%) scale(1); }}
    }}
    #capri_popup {{
        position: fixed; top: 42%; left: 50%;
        transform: translate(-50%, -50%);
        background: #FFFFFF; color: #0D4F66;
        border-top: 5px solid #27AE60;
        border-radius: 14px; padding: 1.6rem 2.4rem;
        box-shadow: 0 12px 40px rgba(0,0,0,0.25);
        z-index: 99999; text-align: center; max-width: 90vw;
        font-size: 1.05rem; font-weight: 600;
        pointer-events: none;
        animation: capri_fade 3.8s ease-in-out forwards;
    }}
    #capri_popup .capri_check {{
        font-size: 0.8rem; color: #27AE60; display: block; margin-bottom: 0.4rem;
        text-transform: uppercase; letter-spacing: 0.12em; font-weight: 700;
    }}
    </style>
    <div id="capri_popup">
        <span class="capri_check">Import réussi</span>
        {_msg_popup}
    </div>
    """, unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
#  ONGLET 1 : ACCUEIL
# ════════════════════════════════════════════════════════════════
with onglet_accueil:
    st.markdown('<div class="section-titre">Bienvenue sur CapriTrack</div>', unsafe_allow_html=True)
    col_texte, col_img = st.columns([2, 1])

    with col_texte:
        st.markdown("""
        **CapriTrack** est un outil de suivi et de visualisation des analyses PCR
        de mycoplasmes caprins, réalisé par les étudiants de **BUT Science des Données**
        dans le cadre d'une SAÉ, pour le compte du **Laboratoire Qualyse**.

        Il permet d'importer les fichiers de résultats PCR, de visualiser l'évolution
        des cheptels session après session, de calculer automatiquement un score
        sanitaire, et d'analyser les tendances pour chacune des 4 cibles de mycoplasmes.
        """)
        st.markdown("#### Tutoriel vidéo")
        st.caption("Vidéo de prise en main de CapriTrack.")
        # Lecteur intégré (la vidéo YouTube doit être en « Non répertoriée »).
        # Pour une démo sans réseau, remplacer l'URL par un fichier local :
        # st.video("assets/tuto.mp4")  (encodé en H.264).
        st.video("https://youtu.be/u1dokiWlK6Y")

    with col_img:
        st.image("assets/equipe.png", use_container_width=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-titre">Comment ça marche ?</div>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    etapes = [
        ("1", "Importez", "Déposez le fichier Excel de résultats PCR (79 ou 86) dans l'onglet Importer."),
        ("2", "Visualisez", "Consultez le tableau pivot, filtrez par cible, département ou session."),
        ("3", "Exportez", "Téléchargez vos résultats au format Excel depuis la visualisation."),
    ]
    for col, (num, titre, desc) in zip([c1, c2, c3], etapes):
        with col:
            st.markdown(f"""
            <div class="etape-card">
                <div class="etape-num">{num}</div>
                <h4 style="color:#0D4F66; margin:0.3rem 0;">{titre}</h4>
                <p style="font-size:0.85rem; color:#5a7a8a;">{desc}</p>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    df_etat = charger_donnees()
    if not df_etat.empty:
        st.markdown('<div class="section-titre">État de la base de données</div>', unsafe_allow_html=True)
        e1, e2, e3, e4 = st.columns(4)
        kpis = [
            (df_etat["num_ede"].nunique(), "Cheptels suivis"),
            (len(df_etat), "Analyses"),
            (df_etat["session_label"].nunique(), "Sessions"),
            (datetime.now().strftime("%d/%m/%Y"), "Dernière consultation"),
        ]
        for col, (val, lab) in zip([e1, e2, e3, e4], kpis):
            with col:
                st.markdown(f'<div class="kpi-card"><div class="kpi-value">{val}</div><div class="kpi-label">{lab}</div></div>', unsafe_allow_html=True)

        # ── Sessions saisies par département ─────────────────────────────────
        st.markdown("<div style='margin-top:2rem'></div>", unsafe_allow_html=True)
        st.markdown("##### Sessions saisies par département")
        st.caption("✓ = au moins un résultat enregistré pour ce département à cette "
                   "session ; : = aucune donnée. Les sessions les plus récentes sont en haut.")
        etat = df_etat.copy()
        etat["Département"] = etat["departement"].map(
            lambda d: "79" if d in ("79", "16") else str(d))
        depts = sorted(etat["Département"].unique())
        # Sessions de la plus récente (en haut) à la plus ancienne (en bas)
        sessions_recent = sorted(etat["session_label"].unique(), reverse=True)
        present = (etat.groupby(["session_label", "Département"]).size()
                   .unstack(fill_value=0) > 0)
        present = present.reindex(index=sessions_recent, columns=depts, fill_value=False)
        disp = present.replace({True: "✓", False: "-"})
        disp.index.name = "Session"

        def _style_presence(v):
            if v == "✓":
                return "color:#0E6E3E; font-weight:700; text-align:center;"
            return "color:#C0392B; text-align:center;"

        sty = disp.style
        sty = (sty.map(_style_presence) if hasattr(sty, "map")
               else sty.applymap(_style_presence))
        sty = sty.set_table_styles([
            {"selector": "th",
             "props": [("background-color", "#0D4F66"), ("color", "white"),
                       ("font-weight", "600"), ("text-align", "center"),
                       ("padding", "6px 16px"), ("font-size", "13px")]},
            {"selector": "td",
             "props": [("padding", "6px 16px"), ("font-size", "14px"),
                       ("border", "1px solid #E0E0E0"), ("text-align", "center")]},
            {"selector": "table", "props": [("border-collapse", "collapse")]},
        ])
        # Toutes les sessions sont affichées ; les plus anciennes sont accessibles
        # au défilement (les plus récentes restent en haut).
        st.markdown(
            f'<div style="max-height:300px; overflow-y:auto; overflow-x:auto;">{sty.to_html()}</div>',
            unsafe_allow_html=True)

    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown(
        "<div style='text-align:center; color:#5a7a8a; font-size:0.85rem; margin-bottom:1.2rem;'>"
        "Un projet réalisé dans le cadre du BUT Science des Données</div>",
        unsafe_allow_html=True,
    )

    # Logo du client : collaboration mise en avant (au-dessus des institutions)
    _qualyse_b64 = _logo_b64("assets/logo_qualyse.png")
    if _qualyse_b64:
        st.markdown(
            f"""<div style='display:flex; flex-direction:column; align-items:center;
            gap:0.55rem; margin-bottom:1.6rem;'>
              <div style='color:#5a7a8a; font-size:0.9rem; letter-spacing:0.02em;'>
              En collaboration avec le laboratoire</div>
              <img src="data:image/png;base64,{_qualyse_b64}" alt="Laboratoire Qualyse"
                   style="height:48px; width:auto; object-fit:contain;">
            </div>""",
            unsafe_allow_html=True,
        )
    # Logos institutionnels : hauteurs individuelles pour un rendu équilibré
    logos_inst = [
        ("assets/logo_poitiers.png", "Université de Poitiers", 80),
        ("assets/logo_iut.png", "IUT Poitiers-Niort-Châtellerault", 120),
        ("assets/logo_butsd.png", "BUT SD Niort", 95),
    ]
    logos_html = ""
    for path, alt, hauteur in logos_inst:
        b64 = _logo_b64(path)
        if b64:
            logos_html += (f'<img src="data:image/png;base64,{b64}" alt="{alt}" '
                           f'style="height:{hauteur}px; width:auto; object-fit:contain;">')
    st.markdown(
        f"""<div style='display:flex; align-items:center; justify-content:space-around;
        gap:2rem; padding:0.5rem 2rem 1.5rem 2rem;'>{logos_html}</div>""",
        unsafe_allow_html=True,
    )


# ════════════════════════════════════════════════════════════════
#  ONGLET 2 : IMPORTER
# ════════════════════════════════════════════════════════════════
with onglet_import:
    st.markdown('<div class="section-titre">Importer un fichier de résultats PCR</div>', unsafe_allow_html=True)

    # ── Session state : évite de re-parser à chaque rerun Streamlit ──────────
    if "import_df"   not in st.session_state: st.session_state.import_df   = None
    if "import_nom"  not in st.session_state: st.session_state.import_nom  = None
    if "import_err"  not in st.session_state: st.session_state.import_err  = None
    # Valeurs de session pré-remplies (modifiables par l'utilisateur ensuite)
    if "import_annee" not in st.session_state: st.session_state.import_annee = datetime.now().year
    if "import_serie" not in st.session_state: st.session_state.import_serie = 1

    # ── 1. Chargement du fichier Excel ────────────────────────────────────────
    st.markdown("#### 1 · Charger le fichier Excel")
    fichier = st.file_uploader("Déposez votre fichier Excel ici", type=["xls", "xlsx"],
                               label_visibility="collapsed",
                               accept_multiple_files=False,
                               key=f"uploader_{st.session_state.get('uploader_cle', 0)}")

    # Reset si l'utilisateur retire le fichier
    if fichier is None:
        st.session_state.import_df  = None
        st.session_state.import_nom = None
        st.session_state.import_err = None
        st.info("Aucun fichier sélectionné pour le moment.")

    else:
        # ── Parse uniquement si c'est un nouveau fichier ──────────────────
        if fichier.name != st.session_state.import_nom:
            st.session_state.import_df  = None
            st.session_state.import_err = None
            suffix = ".xls" if fichier.name.lower().endswith(".xls") else ".xlsx"
            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(fichier.read())
                    tmp_path = tmp.name
                imp = ImportDonnees(tmp_path)
                # Un fichier "mycoïdes générique" (2 cibles) sort de importer() avec
                # une colonne M. MYCOIDES GENERIQUE au lieu des 4 cibles nommées.
                # etendre_colonnes_generiques la déploie en M. MYCOIDES CAPRI +
                # M. CAPRICOLUM + M. PUTREFACIENS : sans cet appel, l'éditeur et la
                # base ne voient jamais M. MYCOIDES CAPRI (cause du bug d'import).
                _df_brut = imp.importer()
                st.session_state.import_df  = imp.etendre_colonnes_generiques(_df_brut)
                st.session_state.import_nom = fichier.name
                # ── Préremplissage auto de la session (point 6) ──────────
                # On utilise le nom RÉEL du fichier (pas le chemin temporaire)
                imp.chemin = fichier.name
                _annee_det, _serie_det = imp._trouver_session()
                if _annee_det is not None:
                    st.session_state.import_annee = int(_annee_det)
                if _serie_det is not None:
                    st.session_state.import_serie = int(_serie_det)
            except Exception as e:
                st.session_state.import_err = str(e)
                st.session_state.import_nom = fichier.name
            finally:
                if tmp_path and os.path.exists(tmp_path):
                    os.unlink(tmp_path)

        # ── Affichage erreur de lecture ───────────────────────────────────
        if st.session_state.import_err:
            st.error(f"Erreur de lecture du fichier : {st.session_state.import_err}")

        df_norm = st.session_state.import_df
        if df_norm is not None:
            st.success(f"Fichier chargé : **{fichier.name}** · "
                       f"{len(df_norm)} lignes · {len(df_norm.columns)} colonnes")

            st.markdown("---")

            # ── 2. Session (pré-remplie automatiquement, modifiable) ──────
            st.markdown("#### 2 · Vérifier la session et le(s) département(s)")
            st.caption("Année et série détectées automatiquement à partir du fichier. "
                       "Corrigez-les si besoin, et vérifiez le(s) département(s) concerné(s).")
            col_an, col_se = st.columns(2)
            with col_an:
                annee = st.number_input(
                    "Année", min_value=2000,
                    max_value=datetime.now().year, step=1,
                    key="import_annee",
                )
            with col_se:
                serie = st.selectbox("Numéro de série", [1, 2, 3], key="import_serie")

            # Départements présents dans le fichier (2 premiers chiffres du n° EDE)
            depts_fichier = sorted({str(e)[:2] for e in df_norm["N EDE"].dropna()})
            options_dept = sorted(set(depts_fichier) | {"79", "86"})
            depts_sel = st.multiselect(
                "Département(s) concerné(s) par cet import",
                options_dept, default=depts_fichier or options_dept,
                key="import_depts",
                help="Une même session (année + série) peut contenir le 79 et le 86, "
                     "importés séparément. L'écrasement ne touchera que le(s) "
                     "département(s) sélectionné(s) ici.")
            if not depts_sel:
                st.info("Sélectionnez au moins un département.")

            # Des données existent-elles DÉJÀ pour cette session ET ces départements ?
            _session_deja_presente = False
            if depts_sel:
                try:
                    _conn_check = sqlite3.connect(DB_PATH)
                    _marq = ",".join("?" * len(depts_sel))
                    _n = _conn_check.execute(
                        f"SELECT COUNT(*) FROM resultat r "
                        f"JOIN prelevement p ON r.id_prelevement = p.id_prelevement "
                        f"JOIN exploitant e ON p.num_ede = e.num_ede "
                        f"WHERE p.annee=? AND p.numero_serie=? AND e.departement IN ({_marq})",
                        (str(int(annee)), int(serie), *depts_sel)
                    ).fetchone()[0]
                    _conn_check.close()
                    _session_deja_presente = _n > 0
                except Exception:
                    _session_deja_presente = False  # base absente/vide : on laisse passer

            _depts_txt = " et ".join(depts_sel) if depts_sel else "-"
            _ecraser = False
            _ecrasement_confirme = False
            if _session_deja_presente:
                st.warning(f"Des résultats existent déjà pour **{int(annee)}_serie{int(serie)}** "
                           f"sur le(s) département(s) **{_depts_txt}**.")
                _ecraser = st.checkbox(
                    "Écraser ces données (remplacer les résultats existants de ce(s) département(s))",
                    key="import_ecraser",
                )
                if _ecraser:
                    st.error(
                        f"Attention : les résultats et prélèvements existants de "
                        f"**{int(annee)}_serie{int(serie)}** pour le(s) département(s) "
                        f"**{_depts_txt}** seront **définitivement supprimés**, puis "
                        "remplacés par le contenu du fichier. Les autres départements "
                        "de cette session ne sont pas touchés.")
                    _ecrasement_confirme = st.checkbox(
                        "Je confirme la suppression définitive de ces données",
                        key="import_confirme_ecrasement",
                    )
                else:
                    st.caption("Sans écrasement, changez d'année/série, ou désélectionnez "
                               "le département déjà présent.")
            elif depts_sel:
                st.success(f"Session **{int(annee)}_serie{int(serie)}** · "
                           f"département(s) {_depts_txt} : prête à être importée.")

            st.markdown("---")

            # ── 3. Prévisualisation éditable ──────────────────────────────
            st.markdown("#### 3 · Vérifier et corriger les données")
            st.caption(
                "Modifiez directement les cellules dans le tableau. Les cellules en "
                "anomalie sont colorées ; survolez-les pour voir le message. "
                "Les corrections sont prises en compte à l'import."
            )

            df_edite, _nb_err, _nb_warn, _nb_info = editeur_donnees_import(df_norm, DB_PATH)

            _bulles = []
            if _nb_err:
                _bulles.append(f"<span style='color:#C0392B;font-weight:700;'>"
                               f"● {_nb_err} erreur(s) à corriger</span>")
            if _nb_warn:
                _bulles.append(f"<span style='color:#E08600;font-weight:700;'>"
                               f"● {_nb_warn} point(s) à vérifier</span>")
            if _nb_info:
                _bulles.append(f"<span style='color:#2F6FB0;font-weight:700;'>"
                               f"● {_nb_info} nouveau(x) cheptel(s)</span>")
            if _bulles:
                st.markdown(
                    "<div style='display:flex;gap:1.5rem;align-items:center;margin:0.4rem 0;flex-wrap:wrap;'>"
                    + "&nbsp;&nbsp;".join(_bulles) + "</div>", unsafe_allow_html=True)
                st.caption("Rouge = erreur bloquante (n° EDE manquant/invalide, Ct aberrant, "
                           "date ou nom manquant). Orange = à vérifier (résultat manquant). "
                           "Bleu = nouveau cheptel (sera créé, non bloquant). Le compteur se "
                           "met à jour à chaque correction ; cliquez sur « Revérifier les "
                           "anomalies » pour rafraîchir les couleurs du tableau.")
            else:
                st.success("Aucune anomalie détectée dans les données.")

            # ── 4. Validation de la structure ─────────────────────────────
            st.markdown("#### 4 · Validation de la structure")
            _v = ImportDonnees.__new__(ImportDonnees)
            _v.df = df_edite
            validation = _v.detecter_tt_col()

            col_ok, col_pcr = st.columns(2)
            with col_ok:
                if validation["manquantes"]:
                    for col in validation["manquantes"]:
                        st.error(f"Colonne manquante : **{col}**")
                else:
                    st.success("Toutes les colonnes obligatoires sont présentes.")
            with col_pcr:
                pcr = validation["pcr"]
                if pcr["valide"]:
                    st.success(f"{pcr['nb']} colonnes PCR détectées.")
                    st.caption(" · ".join(pcr["colonnes"]))
                else:
                    if pcr["nb"] == 0:
                        st.error("Aucune colonne PCR trouvée.")
                    else:
                        st.error(f"{pcr['nb']} colonne(s) PCR détectée(s), alors que 2 ou 4 sont attendues.")
                    st.caption("Attendu : M. MYCOIDES CAPRI · M. CAPRICOLUM · "
                               "M. PUTREFACIENS · M. AGALACTIAE")

            # ── 5. Validation du contenu des données (point 3) ────────────
            st.markdown("#### 5 · Validation des données")
            erreurs_data, avertissements_data = valider_donnees_import(df_edite, DB_PATH)

            if erreurs_data:
                st.error("**Erreurs bloquantes** (à corriger avant l'import) :")
                for msg in erreurs_data:
                    st.markdown(f"- {msg}")
            else:
                st.success("Aucune erreur bloquante détectée dans les données.")

            if avertissements_data:
                with st.container():
                    st.warning("**Avertissements** (n'empêchent pas l'import) :")
                    for msg in avertissements_data:
                        st.markdown(f"- {msg}")

            # ── 6. Lancer l'import ────────────────────────────────────────
            st.markdown("#### 6 · Lancer l'import")
            # La session bloque seulement si des données existent déjà pour ce(s)
            # département(s) ET que l'écrasement n'est pas confirmé.
            _session_bloque = _session_deja_presente and not (_ecraser and _ecrasement_confirme)
            _bloque = (_session_bloque
                       or not depts_sel
                       or not validation["valide"]
                       or len(erreurs_data) > 0)
            if _bloque:
                if not depts_sel:
                    st.warning("Sélectionnez au moins un département (étape 2) avant d'importer.")
                elif _session_bloque:
                    if _ecraser and not _ecrasement_confirme:
                        st.warning("Cochez la case de confirmation (étape 2) pour autoriser l'écrasement.")
                    else:
                        st.warning("Ces données existent déjà pour ce(s) département(s) : "
                                   "cochez « Écraser » (étape 2), ou changez l'année/série/département.")
                elif not validation["valide"]:
                    st.warning("Corrigez les erreurs de structure (étape 4) avant d'importer.")
                else:
                    st.warning("Corrigez les erreurs de données (étape 5) avant d'importer.")
                st.button("Lancer l'import en base", disabled=True,
                          use_container_width=True)
            else:
                _label_btn = ("Écraser et réimporter" if (_ecraser and _ecrasement_confirme)
                              else "Lancer l'import en base")
                if st.button(_label_btn, type="primary", use_container_width=True):
                    with st.spinner("Import en cours…"):
                        _db = DataBase_gestion(DB_PATH)
                        _ok = False
                        _msg = ""
                        _nb_suppr = 0
                        try:
                            _db.connexion_db()
                            # Écrasement éventuel : suppression scopée au(x)
                            # département(s) sélectionné(s) (la sauvegarde reste
                            # à l'initiative de l'utilisateur via l'onglet dédié).
                            if _ecraser and _ecrasement_confirme:
                                _nb_suppr = _db.supprimer_session(
                                    int(annee), int(serie), departements=depts_sel)
                            _session_ok = _db._inp_session(
                                df_edite, annee=int(annee), serie=int(serie)
                            )
                            if not _session_ok:
                                _msg = (f"La session {int(annee)}_serie{int(serie)} "
                                        "n'a pas pu être créée.")
                            else:
                                _db._inp_exploitation(df_edite)
                                _db._inp_matrice(df_edite, auto_accept=True)
                                _db._inp_mycoplasme(df_edite)
                                _db._inp_prelevement(df_edite)
                                _db._inp_resultat(df_edite)
                                _db._inp_clasifier()
                                _db.con.commit()
                                marquer_changement()
                                _ok = True
                                if _nb_suppr > 0:
                                    _msg = (f"Session écrasée et réimportée : "
                                            f"{_nb_suppr} ancien(s) résultat(s) supprimé(s), "
                                            f"{len(df_edite)} ligne(s) importée(s) "
                                            f"pour {int(annee)}_serie{int(serie)}.")
                                else:
                                    _msg = (f"Import réussi : {len(df_edite)} lignes "
                                            f"pour {int(annee)}_serie{int(serie)}.")
                        except Exception as _e:
                            if _db.con:
                                _db.con.rollback()
                            _msg = f"Erreur lors de l'import : {_e}"
                        finally:
                            _db.fermer()

                    if _ok:
                        st.session_state.bdd_modifiee = True
                        charger_donnees.clear()
                        charger_scores.clear()
                        # Popup central de confirmation (affiché après nettoyage)
                        st.session_state.import_succes_msg = _msg
                        # Nettoyage du formulaire : on oublie le fichier chargé et
                        # on réinitialise le file_uploader (via un nouveau key).
                        st.session_state.import_df = None
                        st.session_state.import_nom = None
                        st.session_state.import_err = None
                        st.session_state.uploader_cle = st.session_state.get("uploader_cle", 0) + 1
                        st.rerun()
                    else:
                        st.error(_msg)


# ════════════════════════════════════════════════════════════════
#  ONGLET 3 : VISUALISATION & ANALYSES
# ════════════════════════════════════════════════════════════════
with onglet_visu:
    df = charger_donnees()
    df_scores = charger_scores()

    if df.empty:
        st.warning("La base de données est vide ou introuvable. Importez d'abord des données.")
        st.stop()

    # ── FILTRES (sidebar) ──
    st.sidebar.markdown("### Filtres")
    f_dept = st.sidebar.multiselect("Territoire", sorted(df["zone"].unique()), default=sorted(df["zone"].unique()))
    f_cible = st.sidebar.multiselect("Cible mycoplasme", sorted(df["mycoplasme"].unique()), default=sorted(df["mycoplasme"].unique()))
    sessions_dispo = sorted(df["session_label"].unique())
    f_session = st.sidebar.multiselect("Session", sessions_dispo, default=sessions_dispo)
    f_ede = st.sidebar.text_input("Rechercher un n° EDE")
    st.sidebar.markdown("---")
    st.sidebar.info("CapriTrack · BUT SD\nManon · Louis · Mathéo · Arthur")
    st.sidebar.caption("En collaboration avec le Laboratoire Qualyse")

    # ── APPLICATION DES FILTRES ──
    df_f = df[df["zone"].isin(f_dept) & df["mycoplasme"].isin(f_cible) & df["session_label"].isin(f_session)]
    if f_ede:
        df_f = df_f[df_f["num_ede"].str.contains(f_ede)]

    # ── SOUS-NAVIGATION ──
    sous_onglet = st.radio("Section",
        ["Tableau pivot", "Positivité & négativité", "Classification sanitaire", "Cartographie"],
        horizontal=True, label_visibility="collapsed")

    # ─── A. TABLEAU PIVOT ───
    if sous_onglet == "Tableau pivot":
        st.markdown('<div class="section-titre">Tableau pivot par cible</div>', unsafe_allow_html=True)
        # Le tableau pivot affiche une cible à la fois : on ignore le filtre
        # « cible » global de la barre latérale pour toujours proposer les 4 cibles.
        df_pivot_src = df[df["zone"].isin(f_dept) & df["session_label"].isin(f_session)]
        if f_ede:
            df_pivot_src = df_pivot_src[df_pivot_src["num_ede"].str.contains(f_ede)]
        if df_pivot_src.empty:
            st.warning("Aucune donnée pour ces filtres.")
        else:
            cs1, cs2 = st.columns(2)
            with cs1:
                cible_choisie = st.selectbox("Cible mycoplasme à afficher",
                                             sorted(df_pivot_src["mycoplasme"].unique()))
            with cs2:
                tri_pivot = st.selectbox(
                    "Trier les cheptels par",
                    ["N° EDE (croissant)", "N° EDE (décroissant)",
                     "Moyenne Ct : plus critiques d'abord",
                     "Moyenne Ct : moins critiques d'abord"])
            df_cible = df_pivot_src[df_pivot_src["mycoplasme"] == cible_choisie]

            # ── Filtres locaux du tableau pivot ──
            with st.expander("Filtres du tableau", expanded=False):
                fp1, fp2, fp3 = st.columns(3)
                with fp1:
                    pivot_ede = st.text_input("N° EDE (recherche)", key="pivot_ede", placeholder="Ex : 79345122")
                with fp2:
                    pivot_zones = st.multiselect(
                        "Territoire", sorted(df_cible["zone"].unique()),
                        default=sorted(df_cible["zone"].unique()), key="pivot_zone")
                with fp3:
                    pivot_sessions = st.multiselect(
                        "Session", sorted(df_cible["session_label"].unique()),
                        default=sorted(df_cible["session_label"].unique()), key="pivot_sess")

            # Application des filtres locaux
            df_pivot = df_cible[df_cible["zone"].isin(pivot_zones)]
            if pivot_ede:
                df_pivot = df_pivot[df_pivot["num_ede"].str.contains(pivot_ede)]
            # Filtre session : ne garder que les résultats des sessions choisies
            # (les colonnes non sélectionnées disparaissent, et les cheptels sans
            #  aucun résultat dans ces sessions disparaissent aussi)
            if pivot_sessions:
                df_pivot = df_pivot[df_pivot["session_label"].isin(pivot_sessions)]

            if df_pivot.empty:
                st.warning("Aucun cheptel ne correspond à ces filtres du tableau pivot.")
                pivot_vide = True
            else:
                pivot_vide = False
                pivot = df_pivot.pivot_table(index="num_ede", columns="session_label", values="valeur_ct", aggfunc="first")
                # Tri selon le choix de l'utilisateur
                if tri_pivot == "N° EDE (croissant)":
                    pivot = pivot.sort_index()
                elif tri_pivot == "N° EDE (décroissant)":
                    pivot = pivot.sort_index(ascending=False)
                elif tri_pivot == "Moyenne Ct : moins critiques d'abord":
                    # Ct le plus ÉLEVÉ (le plus négatif/sain) en haut
                    pivot = pivot.reindex(
                        pivot.mean(axis=1, skipna=True)
                             .sort_values(ascending=False, na_position="last").index
                    )
                else:  # Moyenne Ct : plus critiques d'abord (Ct le plus faible en haut)
                    pivot = pivot.reindex(
                        pivot.mean(axis=1, skipna=True)
                             .sort_values(ascending=True, na_position="last").index
                    )
                st.caption(f"{pivot.shape[0]} cheptel(s) · {pivot.shape[1]} session(s) affichée(s).")
                # Rendu maison avec un style en ligne par cellule (voir
                # pivot_vers_html) : garantit la coloration de TOUTES les
                # lignes, y compris les dernières, sur les grands tableaux.
                st.markdown(
                    '<div style="overflow-x:auto; max-height:460px; overflow-y:auto;">'
                    f'{pivot_vers_html(pivot)}</div>',
                    unsafe_allow_html=True,
                )
                st.markdown("""
                <div style="font-size:0.85rem; color:#5a7a8a; display:flex; flex-wrap:wrap; gap:1.2rem; margin-top:0.4rem;">
                    <span><span style="display:inline-block; width:12px; height:12px; background:#D4EFDF; border:1px solid #0E6E3E; border-radius:3px; vertical-align:middle;"></span> Négatif (Ct = 40)</span>
                    <span><span style="display:inline-block; width:12px; height:12px; background:#FDEBD0; border:1px solid #9C5700; border-radius:3px; vertical-align:middle;"></span> Faiblement positif (25 ≤ Ct &lt; 40)</span>
                    <span><span style="display:inline-block; width:12px; height:12px; background:#FADBD8; border:1px solid #922B21; border-radius:3px; vertical-align:middle;"></span> Fortement positif (Ct &lt; 25)</span>
                    <span><span style="display:inline-block; width:12px; height:12px; background:#F0F0F0; border:1px solid #999; border-radius:3px; vertical-align:middle;"></span> Pas de test (NA)</span>
                </div>
                """, unsafe_allow_html=True)

            # Espacement, puis export regroupé dans un volet rattaché au tableau
            st.write("")
            with st.expander("Exporter ce tableau (Excel / PDF)", expanded=False):
                st.caption("Les fichiers Excel et PDF conservent le code couleurs des valeurs Ct.")
                dict_pivots = {}
                for cible in sorted(df["mycoplasme"].unique()):
                    df_c = df[df["mycoplasme"] == cible]
                    _piv = df_c.pivot_table(index="num_ede", columns="session_label",
                                            values="valeur_ct", aggfunc="first")
                    # Modèle Requête complet : trié par N° EDE croissant par défaut
                    _piv = _piv.sort_index()
                    dict_pivots[cible] = _piv
                horodatage = datetime.now().strftime('%Y%m%d')

                st.markdown("**Vue actuelle** (cible sélectionnée, filtres appliqués)")
                if pivot_vide:
                    st.caption("Aucune donnée à exporter pour la vue actuelle (filtres trop restrictifs).")
                else:
                    v1, v2 = st.columns(2)
                    with v1:
                        st.download_button("Excel : vue actuelle",
                            data=exporter_pivots_excel({cible_choisie: pivot}),
                            file_name=f"pivot_{cible_choisie[:20]}_{horodatage}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True)
                    with v2:
                        st.download_button("PDF : vue actuelle",
                            data=exporter_pivots_pdf({cible_choisie: pivot}),
                            file_name=f"pivot_{cible_choisie[:20]}_{horodatage}.pdf",
                            mime="application/pdf", use_container_width=True)

                st.markdown("**Modèle Requête complet** (les 4 cibles)")
                c1, c2 = st.columns(2)
                with c1:
                    st.download_button("Excel : Modèle Requête complet",
                        data=exporter_pivots_excel(dict_pivots),
                        file_name=f"modele_requete_{horodatage}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
                with c2:
                    st.download_button("PDF : Modèle Requête complet",
                        data=exporter_pivots_pdf(dict_pivots),
                        file_name=f"modele_requete_{horodatage}.pdf",
                        mime="application/pdf", use_container_width=True)


    # ─── B. POSITIVITÉ & NÉGATIVITÉ ───
    elif sous_onglet == "Positivité & négativité":
        st.markdown('<div class="section-titre">Positivité &amp; négativité</div>', unsafe_allow_html=True)
        _exp = []

        st.markdown("#### 1. Évolution du nombre de cheptels négatifs")
        st.caption("Nombre de cheptels classés négatifs au fil des sessions "
                   "(toutes cibles confondues).")

        # Plage de sessions affichée pour l'évolution des négatifs.
        # Par défaut : les 3 dernières années présentes.
        sessions_chrono = sorted(df_f["session_label"].unique())
        if len(sessions_chrono) >= 2:
            annees = sorted({s.split("_serie")[0] for s in sessions_chrono})
            annees_def = annees[-3:] if len(annees) > 3 else annees
            sess_def = [s for s in sessions_chrono if s.split("_serie")[0] in annees_def]
            b0, b1 = st.select_slider(
                "Plage de sessions affichée", options=sessions_chrono,
                value=(sess_def[0], sessions_chrono[-1]), key="plage_sessions_desc")
            i0, i1 = sessions_chrono.index(b0), sessions_chrono.index(b1)
            sessions_plage = sessions_chrono[min(i0, i1):max(i0, i1) + 1]
        else:
            sessions_plage = sessions_chrono

        df_neg = (df_f[(df_f["statut"] == "Négatif")
                       & (df_f["session_label"].isin(sessions_plage))]
                  .groupby("session_label")["num_ede"]
                  .nunique().reset_index(name="nb_negatifs").sort_values("session_label"))
        if not df_neg.empty:
            fig1 = px.line(df_neg, x="session_label", y="nb_negatifs", markers=True,
                           labels={"session_label": "Session", "nb_negatifs": "Cheptels négatifs"})
            fig1.update_traces(line_color=COULEURS["turquoise"], line_width=3)
            fig1.update_layout(height=350)
            st.plotly_chart(theme_clair(fig1), use_container_width=True, theme=None)
            _exp.append(("Évolution des cheptels négatifs",
                         df_neg.rename(columns={"session_label": "Session",
                                                "nb_negatifs": "Cheptels négatifs"}),
                         {"type": "line", "titre": "Évolution des cheptels négatifs",
                          "x": "Session", "y": "Cheptels négatifs",
                          "couleur": COULEURS["turquoise"]}))

        st.markdown("#### 2. Évolution du taux de positivité par année")
        st.caption("Part des analyses positives chaque année : la tendance de fond de "
                   "la maladie sur le territoire.")
        df_an = (df_f.groupby("annee")
                 .apply(lambda g: (g["statut"] == "Positif").mean() * 100)
                 .reset_index(name="taux_positif").sort_values("annee"))
        if not df_an.empty:
            df_an["annee"] = df_an["annee"].astype(str)
            fig5 = px.line(df_an, x="annee", y="taux_positif", markers=True,
                           labels={"annee": "Année", "taux_positif": "Taux de positivité (%)"})
            fig5.update_traces(line_color=COULEURS["orange"], line_width=3)
            fig5.update_layout(height=350)
            fig5.update_xaxes(type="category")
            st.plotly_chart(theme_clair(fig5), use_container_width=True, theme=None)
            _exp.append(("Taux par annee",
                              df_an.rename(columns={"annee": "Année",
                                                    "taux_positif": "Taux positivité (%)"}),
                              {"type": "line", "titre": "Taux de positivité par année",
                               "x": "Année", "y": "Taux positivité (%)",
                               "couleur": COULEURS["orange"]}))

        st.markdown("#### 3. Proportion de positifs / négatifs pour une session")
        st.caption("Répartition des cheptels positifs et négatifs pour la session choisie.")
        session_sel = st.selectbox("Session", sessions_dispo)
        df_sess = df_f[df_f["session_label"] == session_sel]
        if not df_sess.empty:
            rep_sess = df_sess["statut"].value_counts().reset_index()
            rep_sess.columns = ["statut", "nb"]
            fig3 = px.pie(rep_sess, names="statut", values="nb", hole=0.5, color="statut",
                          color_discrete_map={"Négatif": COULEURS["turquoise"], "Positif": COULEURS["orange"]})
            fig3.update_layout(height=350)
            st.plotly_chart(theme_clair(fig3), use_container_width=True, theme=None)
            _exp.append((f"Proportion {session_sel}"[:31],
                         rep_sess.rename(columns={"statut": "Statut", "nb": "Nb"}),
                         {"type": "pie", "titre": f"Positifs / négatifs : {session_sel}",
                          "noms": "Statut", "valeurs": "Nb",
                          "couleurs_cat": {"Négatif": COULEURS["turquoise"],
                                           "Positif": COULEURS["orange"]}}))

        st.markdown("#### 4. Effet de la session (1, 2, 3) sur le taux de positifs")
        st.caption("Comparaison du taux de positivité selon la série (1, 2, 3), "
                   "toutes années confondues : met en évidence un éventuel effet saisonnier.")
        df_serie = (df_f.groupby("numero_serie")
                    .apply(lambda g: (g["statut"] == "Positif").mean() * 100)
                    .reset_index(name="taux_positif"))
        if not df_serie.empty:
            df_serie["numero_serie"] = "Série " + df_serie["numero_serie"].astype(str)
            fig4 = px.bar(df_serie, x="numero_serie", y="taux_positif", text="taux_positif",
                          labels={"numero_serie": "Session", "taux_positif": "Taux de positivité (%)"},
                          color_discrete_sequence=[COULEURS["bleu_marine"]])
            fig4.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
            fig4.update_layout(height=350, bargap=0.5)
            fig4.update_xaxes(type="category")
            st.plotly_chart(theme_clair(fig4), use_container_width=True, theme=None)
            _exp.append(("Effet session",
                              df_serie.rename(columns={"numero_serie": "Session",
                                                       "taux_positif": "Taux positivité (%)"}),
                              {"type": "bar", "titre": "Effet de la session sur le taux de positifs",
                               "x": "Session", "y": "Taux positivité (%)",
                               "couleur": COULEURS["bleu_marine"]}))
        # === ZONE À COMPLÉTER : test statistique (chi2/ANOVA) + p-value ===

        st.markdown("#### 5. Comparaison géographique : Deux-Sèvres (79) vs Vienne (86)")
        st.caption("Taux de positivité par territoire. Le département 16 (Charente), "
                   "limitrophe, est rattaché aux Deux-Sèvres.")
        df_dept = (df_f.groupby("zone")
                   .apply(lambda g: (g["statut"] == "Positif").mean() * 100)
                   .reset_index(name="taux_positif"))
        if not df_dept.empty:
            fig6 = px.bar(df_dept, x="zone", y="taux_positif", text="taux_positif",
                          labels={"zone": "Territoire", "taux_positif": "Taux de positivité (%)"},
                          color="zone",
                          color_discrete_map={"Deux-Sèvres (79)": COULEURS["bleu_marine"],
                                              "Vienne (86)": COULEURS["turquoise"]})
            fig6.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
            fig6.update_layout(height=380, showlegend=False, bargap=0.5)
            fig6.update_xaxes(type="category")
            st.plotly_chart(theme_clair(fig6), use_container_width=True, theme=None)
            _exp.append(("Comparaison geo",
                              df_dept.rename(columns={"zone": "Territoire",
                                                      "taux_positif": "Taux positivité (%)"}),
                              {"type": "bar", "titre": "Comparaison géographique 79 vs 86",
                               "x": "Territoire", "y": "Taux positivité (%)",
                               "couleurs_cat": {"Deux-Sèvres (79)": COULEURS["bleu_marine"],
                                                "Vienne (86)": COULEURS["turquoise"]}}))

        _exports_graphiques("posneg", _exp)

    # ─── C. CLASSIFICATION SANITAIRE ───
    elif sous_onglet == "Classification sanitaire":
        st.markdown('<div class="section-titre">Classification sanitaire</div>', unsafe_allow_html=True)
        _exp = []

        # Classes par session (tous mycoplasmes, territoire filtré) : base commune
        # au classement, à la répartition et à l'évolution.
        df_cls = df[df["zone"].isin(f_dept)]
        cps = classes_par_session(df_cls)
        sessions_classees = sorted(cps["session_label"].unique()) if not cps.empty else []

        # ── 1. Classement des cheptels (pour une session) ───────────────────
        st.markdown("#### 1. Classement des cheptels")
        st.caption("Classe sanitaire de chaque cheptel pour la session choisie "
                   "(calculée sur les 6 dernières sessions ≤ 3 ans, toutes cibles confondues).")
        if not sessions_classees:
            st.info("Pas encore assez de sessions pour classer les cheptels "
                    "(6 sessions minimum par cheptel).")
        else:
            sess_classt = st.selectbox("Session", sessions_classees,
                                       index=len(sessions_classees) - 1, key="classt_sess")
            info_cheptel = df_cls.drop_duplicates("num_ede")[["num_ede", "departement", "zone"]]
            df_sc_src = (cps[cps["session_label"] == sess_classt]
                         .merge(info_cheptel, on="num_ede", how="left"))

            # ── Filtres locaux du classement ──
            with st.expander("Filtres du classement", expanded=False):
                fc1, fc2, fc3 = st.columns(3)
                with fc1:
                    classe_ede = st.text_input("N° EDE (recherche)", key="classe_ede",
                                               placeholder="Ex : 79345122")
                with fc2:
                    classe_zones = st.multiselect(
                        "Territoire", sorted(df_sc_src["zone"].dropna().unique()),
                        default=sorted(df_sc_src["zone"].dropna().unique()), key="classe_zone")
                with fc3:
                    classes_dispo = sorted(df_sc_src["classe"].unique())
                    classes_sel = st.multiselect(
                        "Classe", classes_dispo, default=classes_dispo,
                        format_func=lambda c: f"{c} · {LIBELLES_CLASSES.get(c, c)}",
                        key="classe_classe")

            df_sc = df_sc_src[df_sc_src["zone"].isin(classe_zones) &
                              df_sc_src["classe"].isin(classes_sel)].copy()
            if classe_ede:
                df_sc = df_sc[df_sc["num_ede"].str.contains(classe_ede)]

            if df_sc.empty:
                st.info("Aucun cheptel ne correspond à ces filtres.")
            else:
                df_sc["libellé"] = df_sc["classe"].map(LIBELLES_CLASSES)
                df_sc = df_sc.rename(columns={
                    "num_ede": "N° EDE", "departement": "Dépt", "classe": "Classe",
                    "libellé": "Libellé"})
                df_sc = df_sc[["N° EDE", "Dépt", "Classe", "Libellé"]]

                tri_classe = st.selectbox(
                    "Trier par",
                    ["N° EDE (croissant)", "N° EDE (décroissant)",
                     "Classe (A→I)", "Classe (I→A)"], key="tri_classe")
                if tri_classe == "N° EDE (croissant)":
                    df_sc = df_sc.sort_values("N° EDE")
                elif tri_classe == "N° EDE (décroissant)":
                    df_sc = df_sc.sort_values("N° EDE", ascending=False)
                elif tri_classe == "Classe (A→I)":
                    df_sc = df_sc.sort_values(["Classe", "N° EDE"])
                else:
                    df_sc = df_sc.sort_values(["Classe", "N° EDE"], ascending=[False, True])

                st.caption(f"{len(df_sc)} cheptel(s) affiché(s).")

                def _couleur_classe(row):
                    coul = COULEURS_CLASSES.get(row["Classe"], "#FFFFFF")
                    styles = [""] * len(row)
                    idx_classe = list(row.index).index("Classe")
                    styles[idx_classe] = f"background-color:{coul}; color:white; font-weight:700; text-align:center;"
                    return styles

                styler_sc = df_sc.style.apply(_couleur_classe, axis=1).hide(axis="index")
                styler_sc = styler_sc.set_table_styles([
                    {"selector": "th",
                     "props": [("background-color", "#0D4F66"), ("color", "white"),
                               ("font-weight", "600"), ("text-align", "left"),
                               ("padding", "6px 12px"), ("font-size", "13px")]},
                    {"selector": "td",
                     "props": [("padding", "6px 12px"), ("font-size", "13px"),
                               ("border-bottom", "1px solid #E0E0E0"), ("color", "#262730")]},
                    {"selector": "table",
                     "props": [("border-collapse", "collapse"), ("width", "100%"),
                               ("background-color", "white")]},
                ])
                st.markdown(
                    f'<div style="max-height:400px; overflow-y:auto;">{styler_sc.to_html()}</div>',
                    unsafe_allow_html=True)

                st.write("")
                with st.expander("Exporter le classement (Excel / PDF)", expanded=False):
                    he = datetime.now().strftime('%Y%m%d')
                    e1, e2 = st.columns(2)
                    with e1:
                        st.download_button("Excel : classement",
                            data=tableau_simple_excel(df_sc, "Classement",
                                f"Classement des cheptels : {sess_classt}",
                                col_couleur="Classe", couleurs=COULEURS_CLASSES),
                            file_name=f"classement_{sess_classt}_{he}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, key="dl_classt_xlsx")
                    with e2:
                        st.download_button("PDF : classement",
                            data=tableau_simple_pdf(df_sc,
                                f"Classement des cheptels : {sess_classt}",
                                col_couleur="Classe", couleurs=COULEURS_CLASSES),
                            file_name=f"classement_{sess_classt}_{he}.pdf",
                            mime="application/pdf",
                            use_container_width=True, key="dl_classt_pdf")

        # ── 2. Répartition des cheptels par classe (pour une session) ───────
        st.markdown("#### 2. Répartition des cheptels par classe (A à I)")
        st.caption("Nombre de cheptels dans chaque classe sanitaire pour la session "
                   "choisie : photographie du statut sanitaire à un instant donné.")
        if sessions_classees:
            sess_rep = st.selectbox("Session", sessions_classees,
                                    index=len(sessions_classees) - 1, key="rep_classe_sess")
            rep = (cps[cps["session_label"] == sess_rep]
                   .groupby("classe").size().reset_index(name="nb"))
            toutes_classes = pd.DataFrame({"classe": list(LIBELLES_CLASSES.keys())})
            rep = toutes_classes.merge(rep, on="classe", how="left").fillna({"nb": 0})
            rep["nb"] = rep["nb"].astype(int)
            rep["libellé"] = rep["classe"].map(LIBELLES_CLASSES)
            rep = rep.sort_values("classe")
            fig2 = px.bar(rep, x="classe", y="nb", text="nb",
                          labels={"classe": "Classe", "nb": "Nb cheptels"},
                          color="classe", color_discrete_map=COULEURS_CLASSES,
                          hover_data={"libellé": True, "classe": False})
            fig2.update_traces(textposition="outside")
            fig2.update_layout(height=380, showlegend=False)
            fig2.update_xaxes(type="category", categoryorder="array",
                              categoryarray=list(LIBELLES_CLASSES.keys()))
            st.plotly_chart(theme_clair(fig2), use_container_width=True, theme=None)
            _exp.append((f"Répartition classes {sess_rep}"[:31],
                         rep[["classe", "nb"]].rename(columns={"classe": "Classe",
                                                               "nb": "Nb cheptels"}),
                         {"type": "bar", "titre": f"Répartition par classe : {sess_rep}",
                          "x": "Classe", "y": "Nb cheptels",
                          "couleurs_cat": COULEURS_CLASSES}))
        else:
            st.info("Pas encore assez de sessions pour classer les cheptels.")

        # ── 3. Évolution de la répartition des classes (100 %) ──────────────
        st.markdown("#### 3. Évolution de la répartition des classes (100 %)")
        st.caption("Part de chaque classe sanitaire au fil des sessions (barres "
                   "empilées à 100 %). Choisissez ci-dessous la plage de sessions à afficher.")
        if sessions_classees:
            # Plage de sessions propre à ce graphique (défaut : 3 dernières années).
            sessions_chrono = sessions_classees
            if len(sessions_chrono) >= 2:
                annees = sorted({s.split("_serie")[0] for s in sessions_chrono})
                annees_def = annees[-3:] if len(annees) > 3 else annees
                sess_def = [s for s in sessions_chrono if s.split("_serie")[0] in annees_def]
                b0, b1 = st.select_slider("Plage de sessions affichée",
                    options=sessions_chrono, value=(sess_def[0], sessions_chrono[-1]),
                    key="plage_sessions_classif")
                i0, i1 = sessions_chrono.index(b0), sessions_chrono.index(b1)
                sessions_plage = sessions_chrono[min(i0, i1):max(i0, i1) + 1]
            else:
                sessions_plage = sessions_chrono

            cps_p = cps[cps["session_label"].isin(sessions_plage)]
            if not cps_p.empty:
                evol = (cps_p.groupby(["session_label", "classe"]).size()
                        .reset_index(name="nb"))
                evol["pct"] = (evol["nb"]
                               / evol.groupby("session_label")["nb"].transform("sum") * 100)
                fig_ev = px.bar(evol, x="session_label", y="pct", color="classe",
                                category_orders={"classe": list(LIBELLES_CLASSES.keys())},
                                color_discrete_map=COULEURS_CLASSES,
                                labels={"session_label": "Session", "pct": "% des cheptels",
                                        "classe": "Classe"})
                fig_ev.update_layout(barmode="stack", height=420,
                                     yaxis=dict(range=[0, 100], ticksuffix=" %"))
                st.plotly_chart(theme_clair(fig_ev), use_container_width=True, theme=None)
            else:
                st.info("Pas assez de sessions classées dans la plage choisie.")
        else:
            st.info("Pas encore assez de sessions pour tracer l'évolution.")

        _exports_graphiques("classif", _exp)

        # ── Cheptels ayant changé de classe ─────────────────────────────────
        st.markdown("---")
        st.markdown('<div class="section-titre">Cheptels ayant changé de classe</div>',
                    unsafe_allow_html=True)
        st.caption("Choisissez les deux sessions à comparer (par défaut, les deux plus "
                   "récentes). Les cheptels dont la classe a évolué sont à surveiller "
                   "en priorité.")
        _info_cheptels = df_cls.drop_duplicates("num_ede")[
            ["num_ede", "nom", "commune", "departement"]]
        res_chg, s_av_chg, s_ap_chg = afficher_changement_classe(df_cls, _info_cheptels)
        if res_chg is not None and not res_chg.empty:
            with st.expander("Exporter le tableau (Excel / PDF)", expanded=False):
                hc = datetime.now().strftime('%Y%m%d')
                cc1, cc2 = st.columns(2)
                with cc1:
                    st.download_button("Excel : changements",
                        data=tableau_simple_excel(res_chg, "Changements",
                            f"Cheptels ayant changé de classe ({s_av_chg} → {s_ap_chg})"),
                        file_name=f"changements_classe_{hc}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key="dl_chg_xlsx")
                with cc2:
                    st.download_button("PDF : changements",
                        data=tableau_simple_pdf(res_chg,
                            f"Cheptels ayant changé de classe ({s_av_chg} vers {s_ap_chg})"),
                        file_name=f"changements_classe_{hc}.pdf",
                        mime="application/pdf",
                        use_container_width=True, key="dl_chg_pdf")

    # ─── D. CARTOGRAPHIE ───
    elif sous_onglet == "Cartographie":
        afficher_cartographie(df, df_scores)


# ════════════════════════════════════════════════════════════════
#  ONGLET 4 : SAUVEGARDES
# ════════════════════════════════════════════════════════════════
with onglet_sauvegarde:
    st.markdown('<div class="section-titre">Sauvegarde et restauration de la base</div>',
                unsafe_allow_html=True)

    if st.session_state.bdd_modifiee:
        st.warning(
            "Des données ont été importées depuis la dernière sauvegarde. "
            "Pensez à créer une sauvegarde pour garder un point de restauration récent."
        )
    st.caption(
        "Une sauvegarde est automatiquement créée avant chaque écrasement de session. "
        "Vous pouvez aussi en créer une manuellement à tout moment, et restaurer la base "
        "à un état antérieur en cas de besoin. Les 10 sauvegardes les plus récentes sont "
        "conservées."
    )

    if st.button("Créer une sauvegarde maintenant", key="btn_backup_manuel",
                 type="primary", disabled=not st.session_state.bdd_modifiee):
        _chemin = creer_backup()
        if _chemin:
            st.session_state.bdd_modifiee = False
            st.success("Sauvegarde créée.")
        else:
            st.warning("Base introuvable : rien à sauvegarder.")
    if not st.session_state.bdd_modifiee:
        st.caption("Aucune modification depuis la dernière sauvegarde : rien à sauvegarder.")

    st.markdown("---")

    backups = lister_backups()
    if not backups:
        st.info("Aucune sauvegarde pour le moment.")
    else:
        if "backup_voir_plus" not in st.session_state:
            st.session_state.backup_voir_plus = False
        _n = 10 if st.session_state.backup_voir_plus else 3

        st.markdown("**Sauvegardes disponibles** (de la plus récente à la plus ancienne) :")
        for b in backups[:_n]:
            c1, c2 = st.columns([4, 1])
            with c1:
                st.write(f"{b['date_str']} · {b['taille_ko']} Ko")
            with c2:
                if st.button("Restaurer", key=f"restore_{b['nom']}"):
                    st.session_state.backup_a_restaurer = b["chemin"]
                    st.session_state.backup_a_restaurer_date = b["date_str"]

        # Bouton « Afficher plus / moins » seulement s'il y a plus de 3 sauvegardes
        if len(backups) > 3:
            _label = "Afficher moins" if st.session_state.backup_voir_plus else "Afficher plus"
            if st.button(_label, key="btn_backup_voir_plus"):
                st.session_state.backup_voir_plus = not st.session_state.backup_voir_plus
                st.rerun()

        # Confirmation de restauration
        if st.session_state.get("backup_a_restaurer"):
            st.markdown("---")
            st.warning(
                "La restauration va **remplacer la base actuelle** par la sauvegarde "
                f"du **{st.session_state.get('backup_a_restaurer_date', '')}**. "
                "L'état actuel sera lui-même sauvegardé au préalable, par sécurité."
            )
            cc1, cc2 = st.columns(2)
            with cc1:
                if st.button("Confirmer la restauration", type="primary",
                             key="btn_confirme_restore", use_container_width=True):
                    try:
                        restaurer_backup(st.session_state.backup_a_restaurer)
                        charger_donnees.clear()
                        charger_scores.clear()
                        st.session_state.bdd_modifiee = False
                        del st.session_state.backup_a_restaurer
                        st.session_state.pop("backup_a_restaurer_date", None)
                        st.success("Base restaurée.")
                        st.rerun()
                    except Exception as _e:
                        st.error(f"Échec de la restauration : {_e}")
            with cc2:
                if st.button("Annuler", key="btn_annule_restore",
                             use_container_width=True):
                    del st.session_state.backup_a_restaurer
                    st.session_state.pop("backup_a_restaurer_date", None)
                    st.rerun()


# ════════════════════════════════════════════════════════════════
#  AVERTISSEMENT AVANT FERMETURE (si modifications non sauvegardées)
# ════════════════════════════════════════════════════════════════
# Déclenche la fenêtre native du navigateur « Quitter le site ? » lorsque des
# données ont été importées sans qu'une sauvegarde n'ait été créée ensuite.
# Limite technique : le navigateur n'autorise qu'un message générique (texte
# non personnalisable) et ne propose que « Quitter / Rester » : impossible de
# reproduire une vraie boîte « Enregistrer / Ne pas enregistrer » comme Word.
if st.session_state.get("bdd_modifiee"):
    components.html(
        """
        <script>
        const parentWin = window.parent;
        parentWin.onbeforeunload = function (e) {
            e.preventDefault();
            e.returnValue = '';
            return '';
        };
        </script>
        """,
        height=0,
    )
else:
    components.html(
        "<script>window.parent.onbeforeunload = null;</script>",
        height=0,
    )